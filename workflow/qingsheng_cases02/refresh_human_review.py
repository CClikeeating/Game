from __future__ import annotations

import argparse
import json
from pathlib import Path

from .config_loader import OUTPUTS_ROOT, load_config, read_json
from openpyxl import load_workbook

from .pipeline import REVIEW_FIELDS, collect_review_rows, write_human_review


def refresh(batch_id: str) -> dict[str, object]:
    output_dir = OUTPUTS_ROOT / batch_id
    review_rules = load_config("review_rules.yaml")
    previous_edits = existing_user_edits(output_dir / "human_review.xlsx")
    rows = []
    for case_dir in sorted((output_dir / "cases").iterdir()):
        if not case_dir.is_dir():
            continue
        case_card_path = case_dir / "case_card.json"
        if not case_card_path.exists():
            continue
        case_card = read_json(case_card_path)
        rows.extend(collect_review_rows(case_card, review_rules, len(rows) + 1))
    restore_user_edits(rows, previous_edits)
    write_human_review(output_dir / "human_review.xlsx", rows, review_rules)
    return {
        "batch_id": batch_id,
        "human_review": str(output_dir / "human_review.xlsx"),
        "rows": len(rows),
    }


def existing_user_edits(path: Path) -> dict[tuple[str, str, str], dict[str, object]]:
    if not path.exists():
        return {}
    wb = load_workbook(path)
    ws = wb["human_review"]
    headers = [cell.value for cell in ws[1]]
    result = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values)}
        if row.get("your_choice") or row.get("corrected_value") or row.get("notes"):
            key = edit_key(row)
            result[key] = {
                "your_choice": row.get("your_choice", ""),
                "corrected_value": row.get("corrected_value", ""),
                "notes": row.get("notes", ""),
                "status": row.get("status", ""),
            }
    return result


def restore_user_edits(rows: list[dict[str, object]], edits: dict[tuple[str, str, str], dict[str, object]]) -> None:
    for row in rows:
        key = edit_key(row)
        if key in edits:
            row.update(edits[key])


def edit_key(row: dict[str, object]) -> tuple[str, str, str]:
    review_id = str(row.get("review_id") or "")
    if review_id:
        return ("review_id", review_id, "")
    return (
        str(row.get("case_id") or ""),
        str(row.get("field_path") or row.get("field_cn") or ""),
        str(row.get("turn_ids") or ""),
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Refresh human_review.xlsx from existing case_card.json files.")
    parser.add_argument("--batch-id", default="batch_001_data1html_5_cases")
    args = parser.parse_args()
    print(json.dumps(refresh(args.batch_id), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
