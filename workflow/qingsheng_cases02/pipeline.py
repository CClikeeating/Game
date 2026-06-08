from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .config_loader import OUTPUTS_ROOT, load_config, read_json, write_json
from .model_client import ChatModelClient
from .prompt_builder import PRIMARY_SYSTEM_PROMPT, REVIEW_SYSTEM_PROMPT, primary_prompt, review_prompt


ROOT = Path.cwd()
INPUT_BUNDLES_ROOT = ROOT / "outputs" / "source_to_chat_turns01"

REVIEW_FIELDS = [
    "review_id",
    "review_priority",
    "review_scope",
    "review_question",
    "suggested_format",
    "original_path",
    "case_id",
    "field_cn",
    "turn_ids",
    "source_images",
    "source_excerpt",
    "deepseek_value",
    "qwen_value",
    "why_uncertain",
    "impact_if_wrong",
    "your_choice",
    "corrected_value",
    "notes",
    "status",
]

INDEX_FIELDS = [
    "review_id",
    "case_id",
    "source_file",
    "prepared_source_dir",
    "review_type",
    "field_path",
    "field_cn",
    "turn_ids",
    "block_ids",
    "source_images",
]

FIELD_CN = {
    "stage_confidence": "阶段置信度",
    "signals": "女方信号判断",
    "qingsheng_mapping.stage_judgment": "关系阶段综合判断",
    "qingsheng_mapping.stage_judgment.primary_stage": "主判断阶段",
    "qingsheng_mapping.stage_judgment.stage_range": "阶段范围",
    "qingsheng_mapping.stage_judgment.strategy_stage": "策略阶段",
    "qingsheng_mapping.stage_judgment.ambiguity_reason": "阶段模糊原因",
    "qingsheng_mapping.stage_number": "关系阶段编号",
    "qingsheng_mapping.stage_label": "关系阶段标签",
    "qingsheng_mapping.stage_evidence": "阶段判断证据",
    "qingsheng_mapping.cross_stage_signals": "跨阶段穿插信号",
    "qingsheng_mapping.signals": "信号列表",
    "case_facts.male_profile": "男生表现画像",
    "key_moments.good_replies": "男方好回复",
    "key_moments.bad_replies": "男方坏回复",
    "gold_reference.next_reply": "参考下一句",
    "gold_reference.observed_good_reply": "原案例真实好回复",
    "gold_reference.model_suggested_reply": "模型另写回复",
    "gold_reference.reference_type": "参考回复来源",
}


def run_batch(
    batch_id: str,
    input_bundle: str | None = None,
    case_ids: set[str] | None = None,
    case_limit: int | None = None,
    output_batch_id: str | None = None,
    primary_model: str | None = None,
    primary_thinking: str | None = None,
    primary_reasoning_effort: str | None = None,
) -> dict[str, Any]:
    models_config = load_config("models.yaml")
    if primary_model:
        models_config["primary"]["model"] = primary_model
    if primary_thinking:
        models_config["primary"]["thinking"] = {"type": primary_thinking}
    if primary_reasoning_effort:
        models_config["primary"]["reasoning_effort"] = primary_reasoning_effort
    mapping = load_config("qingsheng_mapping.yaml")
    annotation_memory = load_config("annotation_memory.yaml")
    eval_templates = load_config("eval_templates.yaml")
    review_rules = load_config("review_rules.yaml")
    schema_config = load_config("case_schema.yaml")

    input_bundle_dir = resolve_input_bundle(batch_id, input_bundle)
    input_path = input_bundle_dir / "batch_chat_turns.json"
    batch = read_json(input_path)
    output_id = output_batch_id or batch_id
    output_dir = OUTPUTS_ROOT / output_id
    output_dir.mkdir(parents=True, exist_ok=True)

    user_id = str(models_config.get("user_id", "0"))
    primary_client = ChatModelClient("deepseek_primary", models_config["primary"], user_id)
    review_client = ChatModelClient("qwen_review", models_config["review"], user_id)

    manifest_rows: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    model_log: list[dict[str, Any]] = []

    selected_cases = [
        case
        for case in batch.get("cases", [])
        if not case_ids or case.get("case_id", "") in case_ids
    ]
    if case_limit:
        selected_cases = selected_cases[:case_limit]

    for index, case in enumerate(selected_cases, start=1):
        case_id = case["case_id"]
        case_dir = output_dir / "cases" / case_id
        case_dir.mkdir(parents=True, exist_ok=True)

        primary_result = primary_client.chat_json(PRIMARY_SYSTEM_PROMPT, primary_prompt(case, mapping, annotation_memory))
        primary_judgment = primary_result.get("parsed", {})
        review_result = review_client.chat_json(
            REVIEW_SYSTEM_PROMPT,
            review_prompt(case, primary_judgment or fallback_primary_judgment(primary_result), mapping, annotation_memory),
        )
        review_judgment = review_result.get("parsed", {})

        model_log.extend(
            [
                compact_model_log(case_id, "primary", primary_result),
                compact_model_log(case_id, "review", review_result),
            ]
        )

        case_card = build_case_card(
            batch_id=batch_id,
            source_bundle=input_bundle_dir,
            case=case,
            schema_config=schema_config,
            primary_result=primary_result,
            review_result=review_result,
            eval_templates=eval_templates,
        )
        write_case_outputs(case_dir, case_card)

        case_reviews = collect_review_rows(
            case_card=case_card,
            review_rules=review_rules,
            review_start=len(review_rows) + 1,
        )
        review_rows.extend(case_reviews)

        manifest_rows.append(
            {
                "case_id": case_id,
                "source_output": case.get("summary", {}).get("source_output", case_id),
                "turn_count": count_turns(case),
                "primary_model": primary_result.get("model", ""),
                "primary_status": primary_result.get("status", ""),
                "review_model": review_result.get("model", ""),
                "review_status": review_result.get("status", ""),
                "review_item_count": len(case_reviews),
                "status": "needs_human_review" if case_reviews else "ready",
                "case_folder": str(case_dir),
                "case_card_path": str(case_dir / "case_card.json"),
            }
        )
        write_manifest(output_dir, output_id, manifest_rows)
        write_human_review(output_dir / "human_review.xlsx", review_rows, review_rules)
        write_json(output_dir / "model_call_log.json", {"batch_id": batch_id, "output_batch_id": output_id, "calls": model_log})
        print(json.dumps({"case": index, "case_id": case_id, "review_items": len(case_reviews)}, ensure_ascii=False))

    write_manifest(output_dir, output_id, manifest_rows)
    write_human_review(output_dir / "human_review.xlsx", review_rows, review_rules)
    write_json(output_dir / "model_call_log.json", {"batch_id": batch_id, "output_batch_id": output_id, "calls": model_log})
    write_handoff(output_dir, batch_id, output_id, input_bundle_dir)
    return {
        "batch_id": batch_id,
        "output_batch_id": output_id,
        "output_dir": str(output_dir),
        "case_count": len(manifest_rows),
        "selected_case_count": len(selected_cases),
        "review_rows": len(review_rows),
        "primary_model": models_config["primary"]["model"],
        "review_model": models_config["review"]["model"],
        "user_id": user_id,
    }


def resolve_input_bundle(batch_id: str, input_bundle: str | None) -> Path:
    if input_bundle:
        candidate = Path(input_bundle)
        if not candidate.is_absolute():
            candidate = ROOT / candidate
        return candidate
    return INPUT_BUNDLES_ROOT / batch_id


def write_handoff(output_dir: Path, batch_id: str, output_id: str, input_bundle_dir: Path) -> None:
    write_json(
        output_dir / "handoff.json",
        {
            "schema_version": "pipeline_handoff_v1",
            "pipeline": "qingsheng_cases02",
            "batch_id": output_id,
            "source_batch_id": batch_id,
            "source_bundle": str(input_bundle_dir),
            "main_entry": "batch_case_manifest.json",
            "cases_dir": "cases",
            "next_pipeline": "qingsheng_skill_eval03",
            "notes": "Pipeline 3 should read this bundle root.",
        },
    )


def build_case_card(
    batch_id: str,
    source_bundle: Path,
    case: dict[str, Any],
    schema_config: dict[str, Any],
    primary_result: dict[str, Any],
    review_result: dict[str, Any],
    eval_templates: dict[str, Any],
) -> dict[str, Any]:
    primary = primary_result.get("parsed", {})
    review = review_result.get("parsed", {})
    case_id = case.get("case_id", "")
    source_output = case.get("summary", {}).get("source_output", case_id)
    model_agreement = model_agreement_status(primary_result, review_result)
    primary = normalize_primary_judgment(primary)
    eval_cards = build_eval_cards(case, primary, eval_templates)
    quality = build_quality(primary_result, review_result, primary, review, model_agreement)
    return {
        "schema_version": schema_config.get("schema_version", "qingsheng_case_card_v1"),
        "case_meta": {
            "case_id": case_id,
            "batch_id": batch_id,
            "source_output": source_output,
            "source_bundle": str(source_bundle),
            "input_schema_version": "chat_turns_batch_v1",
            "judgment_models": {
                "primary": {
                    "provider": primary_result.get("provider", ""),
                    "model": primary_result.get("model", ""),
                    "user_id": primary_result.get("user_id", ""),
                    "status": primary_result.get("status", ""),
                },
                "review": {
                    "provider": review_result.get("provider", ""),
                    "model": review_result.get("model", ""),
                    "user_id": review_result.get("user_id", ""),
                    "status": review_result.get("status", ""),
                },
            },
        },
        "source_trace": source_trace(case),
        "chat_turns": flatten_turns(case),
        "case_facts": primary.get("case_facts", {}),
        "qingsheng_mapping": primary.get("qingsheng_mapping", {}),
        "key_moments": primary.get("key_moments", {}),
        "gold_reference": primary.get("gold_reference", {}),
        "eval_cards": eval_cards,
        "model_judgments": {
            "primary": primary,
            "review": review,
            "agreement": model_agreement,
        },
        "quality": quality,
    }


def source_trace(case: dict[str, Any]) -> dict[str, Any]:
    images = []
    for block in case.get("blocks", []):
        images.append(
            {
                "block_id": block.get("block_id", ""),
                "source_image": block.get("source_image", ""),
                "crop_box": block.get("crop_box", []),
            }
        )
    return {
        "source_file": case.get("summary", {}).get("source_output", case.get("case_id", "")),
        "block_count": len(case.get("blocks", [])),
        "source_images": images,
    }


def flatten_turns(case: dict[str, Any]) -> list[dict[str, Any]]:
    turns = []
    for block in case.get("blocks", []):
        for turn in block.get("turns", []):
            turns.append(
                {
                    "turn_id": turn.get("turn_id", ""),
                    "block_id": block.get("block_id", turn.get("source_block_id", "")),
                    "speaker": turn.get("speaker", ""),
                    "time": turn.get("time", ""),
                    "text": turn.get("text", ""),
                    "content_type": turn.get("content_type", "text"),
                    "visual_note": turn.get("visual_note", ""),
                    "reason": turn.get("reason", ""),
                    "source_image": turn.get("source_image", block.get("source_image", "")),
                    "need_review": turn.get("need_review", False),
                }
            )
    return turns


def normalize_primary_judgment(primary: dict[str, Any]) -> dict[str, Any]:
    normalized = copy.deepcopy(primary) if isinstance(primary, dict) else {}
    normalize_case_facts(normalized)
    normalize_stage_judgment(normalized)
    gold = normalized.setdefault("gold_reference", {})
    observed = coerce_reply(gold.get("observed_good_reply")) or infer_observed_good_reply(normalized)
    observed = select_best_observed_reply(normalized, observed)
    suggested = str(gold.get("model_suggested_reply") or gold.get("next_reply") or "").strip()

    if observed:
        gold["observed_good_reply"] = observed
        if suggested and suggested != observed["quote"]:
            gold["model_suggested_reply"] = suggested
        else:
            gold.setdefault("model_suggested_reply", "")
        gold["next_reply"] = observed["quote"]
        gold["reference_type"] = "observed_case_reply"
        gold.setdefault(
            "selection_reason",
            "优先使用原案例中真实出现且被判定为有效、可迁移价值更高的男方好回复，不优先使用模型另写的回复。",
        )
    else:
        gold.setdefault("model_suggested_reply", suggested)
        gold["reference_type"] = "model_suggested_reply"
    return normalized


def normalize_case_facts(primary: dict[str, Any]) -> None:
    facts = primary.setdefault("case_facts", {})
    if not isinstance(facts, dict):
        primary["case_facts"] = {}
        facts = primary["case_facts"]
    profile = facts.get("male_profile")
    if not isinstance(profile, dict):
        profile = {}
    facts["male_profile"] = {
        "summary": str(profile.get("summary", "")).strip(),
        "frame_style": str(profile.get("frame_style", "")).strip(),
        "leading_style": str(profile.get("leading_style", "")).strip(),
        "neediness_level": str(profile.get("neediness_level", "")).strip(),
        "communication_traits": profile.get("communication_traits", [])
        if isinstance(profile.get("communication_traits", []), list)
        else [],
        "evidence_turn_ids": profile.get("evidence_turn_ids", [])
        if isinstance(profile.get("evidence_turn_ids", []), list)
        else [],
        "confidence": coerce_float(profile.get("confidence"), 0.0),
        "caveat": str(
            profile.get("caveat")
            or "这是基于当前案例材料的倾向性参考，不是定死判断。"
        ).strip(),
    }


def normalize_stage_judgment(primary: dict[str, Any]) -> None:
    mapping = primary.setdefault("qingsheng_mapping", {})
    if not isinstance(mapping, dict):
        primary["qingsheng_mapping"] = {}
        mapping = primary["qingsheng_mapping"]

    labels = stage_label_map()
    existing = mapping.get("stage_judgment")
    if not isinstance(existing, dict):
        stage_number = coerce_stage_number(mapping.get("stage_number"))
        label = str(mapping.get("stage_label") or labels.get(stage_number, "")).strip()
        confidence = coerce_float(mapping.get("stage_confidence"), 0.0)
        existing = {
            "primary_stage": stage_number,
            "primary_label": label,
            "stage_range": [stage_number, stage_number] if stage_number else [],
            "strategy_stage": stage_number,
            "strategy_label": label,
            "confidence": confidence,
            "ambiguity_reason": "",
            "why_strategy_stage": "",
        }
        mapping["stage_judgment"] = existing

    primary_stage = coerce_stage_number(existing.get("primary_stage"))
    strategy_stage = coerce_stage_number(existing.get("strategy_stage") or primary_stage)
    stage_range = normalize_stage_range(existing.get("stage_range"), primary_stage or strategy_stage)
    confidence = coerce_float(existing.get("confidence", mapping.get("stage_confidence", 0.0)), 0.0)

    existing["primary_stage"] = primary_stage
    existing["primary_label"] = labels.get(primary_stage, "") if primary_stage else ""
    existing["stage_range"] = stage_range
    existing["strategy_stage"] = strategy_stage
    existing["strategy_label"] = labels.get(strategy_stage, "") if strategy_stage else ""
    existing["confidence"] = confidence
    existing.setdefault("ambiguity_reason", "")
    existing.setdefault("why_strategy_stage", "")

    mapping["stage_number"] = strategy_stage
    mapping["stage_label"] = existing["strategy_label"]
    mapping["stage_confidence"] = confidence
    if not isinstance(mapping.get("cross_stage_signals"), list):
        mapping["cross_stage_signals"] = []


def stage_label_map() -> dict[int, str]:
    return {
        1: "阶段1 开场破冰",
        2: "阶段2 建立好感",
        3: "阶段3 关系升温",
        4: "阶段4 邀约见面",
        5: "阶段5 约会实战",
        6: "阶段6 亲密升级",
        7: "阶段7 确立关系",
    }


def coerce_stage_number(value: Any) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return 0
    return number if 1 <= number <= 7 else 0


def coerce_float(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_stage_range(value: Any, fallback_stage: int) -> list[int]:
    stages: list[int] = []
    if isinstance(value, list):
        stages = [coerce_stage_number(item) for item in value]
    elif isinstance(value, str):
        cleaned = value.replace("阶段", "").replace("-", ",").replace("到", ",").replace("至", ",")
        stages = [coerce_stage_number(item.strip()) for item in cleaned.split(",")]
    stages = [stage for stage in stages if stage]
    if not stages and fallback_stage:
        stages = [fallback_stage]
    if len(stages) == 1:
        stages = [stages[0], stages[0]]
    return [min(stages), max(stages)] if stages else []


def infer_observed_good_reply(primary: dict[str, Any]) -> dict[str, str]:
    good_replies = primary.get("key_moments", {}).get("good_replies", []) or []
    candidates = [coerce_reply(item) for item in good_replies]
    candidates = [item for item in candidates if item]
    if not candidates:
        return {}
    return max(candidates, key=reply_score)


def select_best_observed_reply(primary: dict[str, Any], observed: dict[str, str]) -> dict[str, str]:
    good_replies = primary.get("key_moments", {}).get("good_replies", []) or []
    candidates = [coerce_reply(item) for item in good_replies]
    candidates = [item for item in candidates if item]
    if observed:
        candidates.append(observed)
    if not candidates:
        return observed
    best = max(candidates, key=reply_score)
    if not observed:
        return best
    return best if reply_score(best) >= reply_score(observed) + 2 else observed


def reply_score(reply: dict[str, str]) -> int:
    text_blob = " ".join(
        [
            str(reply.get("quote", "")),
            str(reply.get("why_good", "")),
            str(reply.get("transferable_rule", "")),
        ]
    )
    positive_keywords = [
        "框架",
        "化解",
        "测试",
        "后撤",
        "不自证",
        "低需求",
        "边界",
        "排他",
        "没有别人",
        "尊重",
        "稳定",
        "转化",
        "主导",
        "可迁移",
        "邀约",
        "未来",
        "真诚",
        "推进",
        "落地",
        "具体时间",
        "见面",
        "约会",
        "手拉手",
        "牵着",
        "一起",
        "价值观",
        "性格",
        "女朋友",
        "回答我",
        "坚持",
        "正面回答",
    ]
    high_value_quotes = [
        "我们之间没有别人",
        "异地不是问题",
        "感觉才是问题",
        "也不要好人卡",
        "手拉手坐第一排",
        "一起过了才值得",
        "跟你约会",
        "这是找朋友",
        "我挑性格",
        "可你还是没回答我",
    ]
    risky_keywords = [
        "做爱",
        "操",
        "骑",
        "床",
        "胸",
        "亲亲",
        "欲仙欲死",
        "屁屁",
        "屁股",
        "抗揍",
        "拒绝回答",
        "不告诉你",
        "妈妈",
    ]
    score = 0
    for keyword in positive_keywords:
        if keyword in text_blob:
            score += 2
    for keyword in high_value_quotes:
        if keyword in text_blob:
            score += 5
    for keyword in risky_keywords:
        if keyword in text_blob:
            score -= 5
    if reply.get("turn_id"):
        score += 1
    if reply.get("transferable_rule"):
        score += 2
    quote_length = len(str(reply.get("quote", "")))
    if 4 <= quote_length <= 28:
        score += 1
    elif quote_length > 60:
        score -= 1
    return score


def coerce_reply(value: Any) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    quote = str(value.get("quote", "")).strip()
    if not quote:
        return {}
    return {
        "turn_id": str(value.get("turn_id", "")).strip(),
        "quote": quote,
        "why_good": str(value.get("why_good", value.get("why", ""))).strip(),
        "transferable_rule": str(value.get("transferable_rule", "")).strip(),
    }


def build_eval_cards(case: dict[str, Any], primary: dict[str, Any], templates: dict[str, Any]) -> dict[str, Any]:
    transcript = readable_transcript(case)
    base_name = case.get("case_id", "")
    advisory_id = stable_eval_id(base_name, "advisory")
    autopilot_id = stable_eval_id(base_name, "autopilot")
    facts = primary.get("case_facts", {})
    mapping = primary.get("qingsheng_mapping", {})
    stage_judgment = mapping.get("stage_judgment", {}) if isinstance(mapping.get("stage_judgment"), dict) else {}
    male_profile = facts.get("male_profile", {}) if isinstance(facts.get("male_profile"), dict) else {}
    gold = primary.get("gold_reference", {})
    observed = gold.get("observed_good_reply", {}) if isinstance(gold.get("observed_good_reply"), dict) else {}
    rubric = primary.get("eval_rubric", {})
    expected_tail = "\n\n【参考信息给判分人看】\n"
    expected_tail += f"- 模型整理 outcome：{facts.get('outcome', 'unknown')}\n"
    expected_tail += f"- 参考阶段：{stage_summary(mapping)}\n"
    expected_tail += f"- 阶段范围：{json.dumps(stage_judgment.get('stage_range', []), ensure_ascii=False)}\n"
    expected_tail += f"- 阶段模糊原因：{stage_judgment.get('ambiguity_reason', '')}\n"
    expected_tail += f"- 跨阶段穿插信号：{json.dumps(mapping.get('cross_stage_signals', []), ensure_ascii=False)}\n"
    expected_tail += f"- 男生表现画像：{json.dumps(male_profile, ensure_ascii=False)}\n"
    expected_tail += f"- reference type：{gold.get('reference_type', '')}\n"
    expected_tail += f"- 原案例真实好回复：{observed.get('quote', '')}\n"
    expected_tail += f"- gold/reference reply：{gold.get('next_reply', '')}\n"
    expected_tail += f"- advisory must include：{json.dumps(rubric.get('advisory_must_include', []), ensure_ascii=False)}\n"
    expected_tail += f"- must not include：{json.dumps(rubric.get('advisory_must_not_include', []), ensure_ascii=False)}"
    return {
        "advisory": {
            "id": advisory_id,
            "name": f"generated-{base_name}-advisory",
            "prompt": f"{templates['advisory']['prompt_prefix']}\n\n---聊天记录---\n{transcript}\n---聊天记录结束---\n\n轮到我发下一条了。",
            "expected_output": templates["advisory"]["expected_output_template"] + expected_tail,
        },
        "autopilot": {
            "id": autopilot_id,
            "name": f"generated-{base_name}-autopilot",
            "prompt": f"{templates['autopilot']['prompt_prefix']}\n\n---聊天记录---\n{transcript}\n---聊天记录结束---",
            "expected_output": templates["autopilot"]["expected_output_template"] + expected_tail,
        },
    }


def readable_transcript(case: dict[str, Any]) -> str:
    lines = [f"# 对话: {case.get('case_id', '')}"]
    for block in case.get("blocks", []):
        lines.append(f"\n--- block: {block.get('block_id', '')} ---")
        for turn in block.get("turns", []):
            time = f"[{turn.get('time')}] " if turn.get("time") else ""
            lines.append(f"{time}{speaker_cn(turn.get('speaker', 'unknown'))}：{turn.get('text', '')}")
    return "\n".join(lines)


def stage_summary(mapping: dict[str, Any]) -> str:
    judgment = mapping.get("stage_judgment", {}) if isinstance(mapping.get("stage_judgment"), dict) else {}
    strategy_label = judgment.get("strategy_label") or mapping.get("stage_label", "")
    stage_range = judgment.get("stage_range", [])
    if isinstance(stage_range, list) and len(stage_range) >= 2 and stage_range[0] != stage_range[-1]:
        return f"{strategy_label}（范围：阶段{stage_range[0]}-阶段{stage_range[-1]}）"
    return str(strategy_label)


def speaker_cn(speaker: str) -> str:
    return {
        "male": "男",
        "female": "女",
        "system": "系统",
        "narration": "旁白",
    }.get(speaker, "未知")


def stable_eval_id(case_id: str, mode: str) -> int:
    digest = hashlib.sha1(f"{case_id}:{mode}".encode("utf-8")).hexdigest()
    return 100000 + int(digest[:8], 16) % 800000


def build_quality(
    primary_result: dict[str, Any],
    review_result: dict[str, Any],
    primary: dict[str, Any],
    review: dict[str, Any],
    model_agreement: str,
) -> dict[str, Any]:
    issues = []
    if primary_result.get("status") != "model_success":
        issues.append({"type": "primary_model_failed", "detail": primary_result.get("error", "")})
    if review_result.get("status") != "model_success":
        issues.append({"type": "review_model_failed", "detail": review_result.get("error", "")})
    for item in primary.get("quality", {}).get("uncertain_items", []) or []:
        issues.append({"type": "primary_uncertain", **item})
    for item in review.get("additional_uncertain_items", []) or []:
        issues.append({"type": "review_uncertain", **item})
    for conflict in review.get("conflicts", []) or []:
        issues.append({"type": "model_conflict", **conflict})
    mapping = primary.get("qingsheng_mapping", {})
    stage_judgment = mapping.get("stage_judgment", {}) if isinstance(mapping.get("stage_judgment"), dict) else {}
    confidence = stage_judgment.get("confidence", mapping.get("stage_confidence"))
    if isinstance(confidence, (int, float)) and confidence < 0.72:
        issues.append(
            {
                "type": "low_stage_confidence",
                "field": "qingsheng_mapping.stage_judgment.confidence",
                "primary_value": confidence,
                "impact": "阶段判断会影响 skill 的回复策略和 eval 标准",
            }
        )
    if isinstance(stage_judgment, dict) and not stage_judgment.get("stage_range"):
        issues.append(
            {
                "type": "missing_stage_range",
                "field": "qingsheng_mapping.stage_judgment.stage_range",
                "primary_value": stage_judgment,
                "impact": "缺少阶段范围会让边界模糊案例被硬判为单点阶段",
            }
        )
    if isinstance(stage_judgment, dict):
        stage_range = stage_judgment.get("stage_range", [])
        if isinstance(stage_range, list) and len(stage_range) >= 2:
            try:
                stage_width = int(stage_range[-1]) - int(stage_range[0])
            except (TypeError, ValueError):
                stage_width = 0
            if stage_width > 3:
                issues.append(
                    {
                        "type": "overwide_stage_range",
                        "field": "qingsheng_mapping.stage_judgment.stage_range",
                        "primary_value": stage_range,
                        "reason": "阶段范围超过3个连续阶段，可能被成功结果倒推过头",
                        "impact": "阶段跨度过宽会让 skill/eval 的策略标准失焦，需要人工确认主阶段和策略阶段",
                    }
                )
    if mapping.get("stage_number") == 7 and not has_stage7_evidence(mapping):
        issues.append(
            {
                "type": "stage7_evidence_check",
                "field": "qingsheng_mapping.stage_number",
                "primary_value": mapping.get("stage_number"),
                "reason": "判为阶段7但未看到明确关系确认、承诺或排他关系证据",
                "impact": "把调情或亲密想象误判为确立关系，会导致后续策略过激进",
            }
        )
    need_review = bool(issues) or model_agreement != "model_agreed"
    return {
        "model_agreement": model_agreement,
        "need_human_review": need_review,
        "review_items": issues,
    }


def model_agreement_status(primary_result: dict[str, Any], review_result: dict[str, Any]) -> str:
    if primary_result.get("status") != "model_success" or review_result.get("status") != "model_success":
        return "model_incomplete"
    verdict = str(review_result.get("parsed", {}).get("verdict", "")).lower()
    if verdict == "agree":
        return "model_agreed"
    if verdict in {"partial", "disagree"}:
        return "model_disagreed"
    return "review_unclear"


def has_stage7_evidence(mapping: dict[str, Any]) -> bool:
    evidence = mapping.get("stage_evidence", []) or []
    cross_signals = mapping.get("cross_stage_signals", []) or []
    text = json.dumps([evidence, cross_signals], ensure_ascii=False)
    strong_keywords = ["确定关系", "确立关系", "在一起", "女朋友", "男朋友", "对象", "正式关系", "确认关系"]
    weak_only_keywords = ["老公", "老婆", "抱抱", "亲", "睡觉", "做爱", "想你"]
    if any(keyword in text for keyword in strong_keywords):
        return True
    if any(keyword in text for keyword in weak_only_keywords):
        return False
    return False


def collect_review_rows(
    case_card: dict[str, Any],
    review_rules: dict[str, Any],
    review_start: int,
) -> list[dict[str, Any]]:
    rows = []
    case_id = case_card["case_meta"]["case_id"]
    source_file = case_card["source_trace"]["source_file"]
    source_map = source_location_map(case_card)
    pending_items = [
        item
        for item in (case_card.get("quality", {}).get("review_items", []) or [])
        if not item.get("human_review_status")
    ]
    for offset, item in enumerate(group_review_items(pending_items)):
        turn_ids = item.get("evidence_turn_ids") or [item.get("turn_id", "")]
        turn_ids = [turn_id for turn_id in turn_ids if turn_id]
        turn_ids = review_turn_ids(item, case_card, turn_ids)
        block_ids = block_ids_for_turns(case_card, turn_ids)
        source_images = source_images_for_blocks(case_card, block_ids)
        rows.append(
            {
                "review_id": f"review_{review_start + offset:04d}",
                "review_priority": review_priority(item),
                "review_scope": review_scope(item, case_card, turn_ids),
                "review_question": review_question(item, case_card),
                "suggested_format": review_suggested_format(item),
                "case_id": case_id,
                "source_file": source_file,
                "original_path": source_map.get("original_path", ""),
                "prepared_source_dir": source_map.get("prepared_source_dir", ""),
                "review_type": item.get("type", "uncertain"),
                "field_path": item.get("field", ""),
                "field_cn": field_path_cn(str(item.get("field", ""))),
                "turn_ids": ", ".join(turn_ids),
                "block_ids": ", ".join(block_ids),
                "source_images": "\n".join(source_images),
                "source_excerpt": source_excerpt(case_card, turn_ids),
                "deepseek_value": model_side_value(case_card, item, "primary"),
                "qwen_value": model_side_value(case_card, item, "review"),
                "why_uncertain": item.get("reason", item.get("detail", "")),
                "impact_if_wrong": item.get("impact", ""),
                "your_choice": "",
                "corrected_value": "",
                "notes": "",
                "status": "pending",
            }
        )
    return rows


def review_turn_ids(item: dict[str, Any], case_card: dict[str, Any], turn_ids: list[str]) -> list[str]:
    if turn_ids:
        return turn_ids
    field = str(item.get("field", ""))
    item_type = str(item.get("type", ""))
    mapping = case_card.get("qingsheng_mapping", {})
    if is_stage_review(field, item_type):
        evidence_turns = []
        for evidence in mapping.get("stage_evidence", []) or []:
            turn_id = str(evidence.get("turn_id", "")).strip()
            if turn_id:
                evidence_turns.append(turn_id)
        for signal in mapping.get("cross_stage_signals", []) or []:
            turn_id = str(signal.get("turn_id", "")).strip()
            if turn_id:
                evidence_turns.append(turn_id)
        return list(dict.fromkeys(evidence_turns))[:8]
    if "gold_reference" in field:
        observed = case_card.get("gold_reference", {}).get("observed_good_reply", {})
        if isinstance(observed, dict) and observed.get("turn_id"):
            return [str(observed["turn_id"])]
    return []


def is_stage_review(field: str, item_type: str) -> bool:
    return (
        "stage" in field
        or item_type
        in {
            "overwide_stage_range",
            "stage7_evidence_check",
            "low_stage_confidence",
            "missing_stage_range",
        }
    )


def review_scope(item: dict[str, Any], case_card: dict[str, Any], turn_ids: list[str]) -> str:
    field = str(item.get("field", ""))
    item_type = str(item.get("type", ""))
    if is_stage_review(field, item_type):
        mapping = case_card.get("qingsheng_mapping", {})
        stage_judgment = mapping.get("stage_judgment", {}) if isinstance(mapping.get("stage_judgment"), dict) else {}
        return (
            "判断范围：这组阶段证据/当前策略判断点，不是给整套完整聊天只定一个阶段。"
            f"当前主阶段={stage_judgment.get('primary_stage')}，"
            f"策略阶段={stage_judgment.get('strategy_stage')}，"
            f"阶段范围={stage_judgment.get('stage_range')}。"
            f"重点参考 turn_ids：{', '.join(turn_ids) if turn_ids else '模型未给出证据，请结合source_excerpt/原文件判断'}。"
            "如果这些turn横跨多个小阶段，请判断当前最关键片段应按哪个阶段处理。"
        )
    if "gold_reference" in field:
        return "判断范围：gold/reference reply 是否适合作为本案例可学习的男方好回复。重点看对应 turn 和前后上下文。"
    if "male_profile" in field:
        return "判断范围：男生在本案例里的表现画像，不是判断这个人的固定人格。"
    return "判断范围：这一条模型不确定或模型冲突对应的字段。优先看 source_excerpt，再看原图/原文件。"


def review_priority(item: dict[str, Any]) -> str:
    item_type = str(item.get("type", ""))
    field = str(item.get("field", ""))
    if item_type in {"primary_model_failed", "review_model_failed", "stage7_evidence_check"}:
        return "must_review"
    if "gold_reference" in field or item_type in {"overwide_stage_range", "missing_stage_range"}:
        return "must_review"
    if item_type == "model_conflict":
        return "must_review"
    if item_type in {"low_stage_confidence", "model_uncertain"}:
        return "optional_review"
    return "optional_review"


def review_question(item: dict[str, Any], case_card: dict[str, Any]) -> str:
    item_type = str(item.get("type", ""))
    field = str(item.get("field", ""))
    mapping = case_card.get("qingsheng_mapping", {})
    stage_judgment = mapping.get("stage_judgment", {}) if isinstance(mapping.get("stage_judgment"), dict) else {}
    if "stage_judgment" in field or field in {"qingsheng_mapping.stage_number", "qingsheng_mapping.stage_label"}:
        return (
            "请判断这组阶段证据/当前策略判断点的阶段是否合理，不是给整套聊天只定一个阶段。当前："
            f"主阶段={stage_judgment.get('primary_stage')}，"
            f"策略阶段={stage_judgment.get('strategy_stage')}，"
            f"阶段范围={stage_judgment.get('stage_range')}。"
            "如果范围过宽，请在 corrected_value 写如 [3,4]；如果阶段错了，写这个判断点的主阶段/策略阶段建议。"
        )
    if item_type == "overwide_stage_range":
        return (
            f"模型给出的阶段范围 {item.get('primary_value')} 可能过宽。"
            "请根据 review_scope 里的阶段证据和 source_excerpt 判断应缩窄到哪个范围，例如 [3,4] 或 [4,5]。"
        )
    if item_type == "stage7_evidence_check":
        return "模型可能把亲密调情误判为阶段7。请确认是否真的已经确立关系；如果没有，请写更合适的阶段。"
    if "gold_reference" in field:
        return "请判断 gold/reference reply 是否应该作为案例标准好回复。若不合适，请在 corrected_value 写更合适的原句或 turn_id。"
    if "male_profile" in field:
        return "请判断男生表现画像是否过度、太绝对或证据不足。可写更准确的画像描述。"
    if "cross_stage_signals" in field:
        return "请判断这是否只是局部跨阶段信号，而不是整体主阶段。可写应保留、删除或调整为哪个阶段信号。"
    if item_type == "model_conflict":
        return "DeepSeek 和 Qwen 判断不一致。请根据上下文选择更合理的一方，或手动写最终判断。"
    if item_type in {"primary_model_failed", "review_model_failed"}:
        return "模型调用失败或输出异常。请人工给出该字段最终判断，或标记为不确定。"
    return "请根据 source_excerpt 和两边模型意见判断是否需要修正。若同意某一方，直接在 your_choice 选择即可。"


def review_suggested_format(item: dict[str, Any]) -> str:
    field = str(item.get("field", ""))
    item_type = str(item.get("type", ""))
    if "stage_range" in field or item_type == "overwide_stage_range":
        return "示例：[3,4]"
    if "stage_number" in field or "stage_judgment" in field or item_type == "stage7_evidence_check":
        return "示例：主阶段3，策略阶段4，范围[3,4]"
    if "gold_reference" in field:
        return "示例：turn_0123：原句；或直接写正确原句"
    if "male_profile" in field:
        return "示例：框架较强，主动引导，需求感中等；证据 turn_xxx"
    return "可留空；需要修正时写中文说明或 JSON"


def field_path_cn(field_path: str) -> str:
    if field_path in FIELD_CN:
        return FIELD_CN[field_path]
    cleaned = field_path
    cleaned = cleaned.replace("qingsheng_mapping.", "")
    cleaned = cleaned.replace("key_moments.", "")
    cleaned = cleaned.replace("case_facts.", "")
    cleaned = cleaned.replace("gold_reference.", "")
    cleaned = cleaned.replace("stage_judgment", "阶段综合判断")
    cleaned = cleaned.replace("primary_stage", "主判断阶段")
    cleaned = cleaned.replace("primary_label", "主判断阶段标签")
    cleaned = cleaned.replace("stage_range", "阶段范围")
    cleaned = cleaned.replace("strategy_stage", "策略阶段")
    cleaned = cleaned.replace("strategy_label", "策略阶段标签")
    cleaned = cleaned.replace("ambiguity_reason", "阶段模糊原因")
    cleaned = cleaned.replace("why_strategy_stage", "策略阶段理由")
    cleaned = cleaned.replace("confidence", "置信度")
    cleaned = cleaned.replace("stage_evidence", "阶段证据")
    cleaned = cleaned.replace("cross_stage_signals", "跨阶段穿插信号")
    cleaned = cleaned.replace("stage_number", "阶段编号")
    cleaned = cleaned.replace("stage_label", "阶段标签")
    cleaned = cleaned.replace("stage_confidence", "阶段置信度")
    cleaned = cleaned.replace("signals", "信号")
    cleaned = cleaned.replace("male_profile", "男生表现画像")
    cleaned = cleaned.replace("good_replies", "好回复")
    cleaned = cleaned.replace("bad_replies", "坏回复")
    cleaned = cleaned.replace("turning_points", "关键转折")
    cleaned = cleaned.replace("next_reply", "参考下一句")
    return cleaned or "未指定字段"


def group_review_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    output: list[dict[str, Any]] = []
    for item in items:
        item_type = item.get("type", "")
        if item_type not in {"primary_uncertain", "review_uncertain"}:
            output.append(item)
            continue
        key = (
            str(item.get("field", "")),
            ",".join(item.get("evidence_turn_ids") or [str(item.get("turn_id", ""))]),
        )
        group = grouped.setdefault(
            key,
            {
                "type": "model_uncertain",
                "field": item.get("field", ""),
                "turn_id": item.get("turn_id", ""),
                "evidence_turn_ids": item.get("evidence_turn_ids", []),
                "primary_item": None,
                "review_item": None,
                "reason": "",
                "impact": "",
            },
        )
        if item.get("evidence_turn_ids") and not group.get("evidence_turn_ids"):
            group["evidence_turn_ids"] = item.get("evidence_turn_ids", [])
        if item.get("turn_id") and not group.get("turn_id"):
            group["turn_id"] = item.get("turn_id")
        if item_type == "primary_uncertain":
            group["primary_item"] = item
        else:
            group["review_item"] = item
        group["reason"] = merge_texts(
            str(group.get("reason", "")),
            f"{'DeepSeek' if item_type == 'primary_uncertain' else 'Qwen'}：{item.get('reason', '')}",
        )
        group["impact"] = merge_texts(str(group.get("impact", "")), str(item.get("impact", "")))
    output.extend(grouped.values())
    return output


def merge_texts(left: str, right: str) -> str:
    right = right.strip()
    if not right:
        return left
    if right in left:
        return left
    return f"{left}\n{right}".strip()


def source_location_map(case_card: dict[str, Any]) -> dict[str, str]:
    case_id = case_card.get("case_meta", {}).get("case_id", "")
    source_bundle = Path(str(case_card.get("case_meta", {}).get("source_bundle", "")))
    prepared_dir = source_bundle / "cases" / case_id if source_bundle else Path()
    manifest_path = prepared_dir / "source_manifest.json"
    original_path = ""
    if manifest_path.exists():
        manifest = read_json(manifest_path)
        original_path = str(manifest.get("original_path", ""))
    return {
        "original_path": original_path,
        "prepared_source_dir": str(prepared_dir),
    }


def block_ids_for_turns(case_card: dict[str, Any], turn_ids: list[str]) -> list[str]:
    if not turn_ids:
        return []
    by_turn = {turn.get("turn_id", ""): turn.get("block_id", "") for turn in case_card.get("chat_turns", [])}
    return list(dict.fromkeys(block for turn_id in turn_ids for block in [by_turn.get(turn_id, "")] if block))


def source_images_for_blocks(case_card: dict[str, Any], block_ids: list[str]) -> list[str]:
    if not block_ids:
        return []
    by_block = {
        item.get("block_id", ""): item.get("source_image", "")
        for item in case_card.get("source_trace", {}).get("source_images", [])
    }
    return list(dict.fromkeys(str(by_block.get(block_id, "")) for block_id in block_ids if by_block.get(block_id, "")))


def model_side_value(case_card: dict[str, Any], item: dict[str, Any], side: str) -> str:
    if item.get("type") == "model_uncertain":
        nested = item.get("primary_item") if side == "primary" else item.get("review_item")
        if isinstance(nested, dict):
            return analysis_summary("DeepSeek" if side == "primary" else "Qwen", nested)
        return no_separate_opinion(case_card, item, side)
    if side == "primary":
        direct = item.get("primary_value", "")
        if direct not in ("", None):
            return stringify_value(direct)
        if item.get("type") == "primary_uncertain":
            return analysis_summary("DeepSeek", item)
        if item.get("type") in {"review_uncertain", "model_conflict"}:
            path_value = value_by_path(case_card.get("model_judgments", {}).get("primary", {}), str(item.get("field", "")))
            if path_value not in ("", None):
                return stringify_value(path_value)
        return analysis_summary("DeepSeek", item)
    direct = item.get("review_value", "")
    if direct not in ("", None):
        return stringify_value(direct)
    if item.get("type") == "review_uncertain":
        return analysis_summary("Qwen", item)
    if item.get("type") in {"primary_uncertain", "model_conflict"}:
        path_value = value_by_path(case_card.get("model_judgments", {}).get("review", {}), str(item.get("field", "")))
        if path_value not in ("", None):
            return stringify_value(path_value)
    return analysis_summary("Qwen", item)


def no_separate_opinion(case_card: dict[str, Any], item: dict[str, Any], side: str) -> str:
    if side == "primary":
        path_value = value_by_path(case_card.get("model_judgments", {}).get("primary", {}), str(item.get("field", "")))
        parts = ["DeepSeek未单独提出这个复核疑点。"]
        if path_value not in ("", None):
            parts.append(f"DeepSeek对应字段值：{stringify_value(path_value)}")
        return "\n".join(parts)
    review = case_card.get("model_judgments", {}).get("review", {})
    notes = review.get("quality", {}).get("notes", "")
    verdict = review.get("verdict", "")
    parts = ["Qwen未单独提出这个复核疑点。"]
    if verdict:
        parts.append(f"Qwen总体复核结论：{verdict}")
    if notes:
        parts.append(f"Qwen总体说明：{notes}")
    return "\n".join(parts)


def analysis_summary(model_name: str, item: dict[str, Any]) -> str:
    parts = []
    reason = item.get("reason", item.get("detail", ""))
    impact = item.get("impact", "")
    if reason:
        parts.append(f"{model_name}分析：{reason}")
    if impact:
        parts.append(f"影响：{impact}")
    return "\n".join(parts)


def value_by_path(data: Any, field_path: str) -> Any:
    if not field_path:
        return ""
    current = data
    for raw_part in field_path.split("."):
        part = raw_part.strip()
        if not part:
            continue
        if "[" in part and part.endswith("]"):
            name, index_text = part[:-1].split("[", 1)
            current = current.get(name, []) if isinstance(current, dict) else []
            if not index_text.isdigit() or int(index_text) >= len(current):
                return ""
            current = current[int(index_text)]
            continue
        if isinstance(current, dict):
            current = current.get(part, "")
        else:
            return ""
    return current


def source_excerpt(case_card: dict[str, Any], turn_ids: list[str]) -> str:
    if not turn_ids:
        return ""
    turns = case_card.get("chat_turns", [])
    selected = []
    for index, turn in enumerate(turns):
        if turn.get("turn_id") in turn_ids:
            start = max(0, index - 2)
            end = min(len(turns), index + 3)
            for item in turns[start:end]:
                selected.append(f"{item.get('turn_id')} {speaker_cn(item.get('speaker', ''))}：{item.get('text', '')}")
    return "\n".join(dict.fromkeys(selected))


def stringify_value(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def fallback_primary_judgment(result: dict[str, Any]) -> dict[str, Any]:
    return {
        "quality": {
            "need_human_review": True,
            "uncertain_items": [
                {
                    "field": "model.primary",
                    "reason": result.get("error", result.get("status", "")),
                    "impact": "主判断模型失败，案例判断需要人工或复核模型确认",
                }
            ],
        }
    }


def compact_model_log(case_id: str, role: str, result: dict[str, Any]) -> dict[str, Any]:
    return {
        "case_id": case_id,
        "role": role,
        "provider": result.get("provider", ""),
        "model": result.get("model", ""),
        "user_id": result.get("user_id", ""),
        "status": result.get("status", ""),
        "error": result.get("error", ""),
        "elapsed_seconds": result.get("elapsed_seconds", 0),
        "usage": result.get("usage", {}),
        "raw_text_preview": str(result.get("raw_text", ""))[:1200],
    }


def write_case_outputs(case_dir: Path, case_card: dict[str, Any]) -> None:
    write_json(case_dir / "case_card.json", case_card)
    write_json(case_dir / "eval_advisory.json", case_card["eval_cards"]["advisory"])
    write_json(case_dir / "eval_autopilot.json", case_card["eval_cards"]["autopilot"])
    write_json(case_dir / "case_quality_report.json", case_card["quality"])
    (case_dir / "readable_case.md").write_text(build_readable_case(case_card), encoding="utf-8")


def build_readable_case(case_card: dict[str, Any]) -> str:
    facts = case_card.get("case_facts", {})
    mapping = case_card.get("qingsheng_mapping", {})
    stage_judgment = mapping.get("stage_judgment", {}) if isinstance(mapping.get("stage_judgment"), dict) else {}
    male_profile = facts.get("male_profile", {}) if isinstance(facts.get("male_profile"), dict) else {}
    gold = case_card.get("gold_reference", {})
    observed = gold.get("observed_good_reply", {}) if isinstance(gold.get("observed_good_reply"), dict) else {}
    def text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    lines = [
        f"# {case_card['case_meta']['case_id']}",
        "",
        "## 案例概况",
        f"- 关系弧线：{text(facts.get('relationship_arc', ''))}",
        f"- 男方目标：{text(facts.get('male_goal', ''))}",
        f"- 女方状态：{text(facts.get('female_state', ''))}",
        f"- 最终走向：{text(facts.get('outcome', ''))}",
        "",
        "## 男生表现画像",
        f"- 概括：{text(male_profile.get('summary', ''))}",
        f"- 框架风格：{text(male_profile.get('frame_style', ''))}",
        f"- 引导方式：{text(male_profile.get('leading_style', ''))}",
        f"- 需求感倾向：{text(male_profile.get('neediness_level', ''))}",
        f"- 沟通特质：{text(male_profile.get('communication_traits', []))}",
        f"- 依据 turn_id：{text(male_profile.get('evidence_turn_ids', []))}",
        f"- 置信度：{text(male_profile.get('confidence', ''))}",
        f"- 注意：{text(male_profile.get('caveat', ''))}",
        "",
        "## qingsheng 映射",
        f"- 策略阶段：{stage_summary(mapping)}",
        f"- 主判断阶段：{text(stage_judgment.get('primary_label', ''))}",
        f"- 阶段范围：{text(stage_judgment.get('stage_range', ''))}",
        f"- 置信度：{text(stage_judgment.get('confidence', mapping.get('stage_confidence', '')))}",
        f"- 模糊原因：{text(stage_judgment.get('ambiguity_reason', ''))}",
        f"- 策略阶段理由：{text(stage_judgment.get('why_strategy_stage', ''))}",
        f"- 跨阶段穿插信号：{text(mapping.get('cross_stage_signals', []))}",
        "",
        "## 原案例真实好回复",
        f"- turn_id：{text(observed.get('turn_id', ''))}",
        f"- 原句：{text(observed.get('quote', ''))}",
        f"- 可迁移规则：{text(observed.get('transferable_rule', ''))}",
        "",
        "## 参考下一句",
        text(gold.get("next_reply", "")),
        "",
        "## 模型另写回复",
        text(gold.get("model_suggested_reply", "")),
        "",
        "## 需要复核",
        f"- {case_card.get('quality', {}).get('need_human_review', False)}",
        "",
        "## 对话",
    ]
    for turn in case_card.get("chat_turns", []):
        lines.append(f"- `{text(turn.get('turn_id'))}` {speaker_cn(turn.get('speaker', ''))}：{text(turn.get('text', ''))}")
    return "\n".join(lines)


def write_manifest(output_dir: Path, batch_id: str, rows: list[dict[str, Any]]) -> None:
    write_json(output_dir / "batch_case_manifest.json", {"batch_id": batch_id, "cases": rows})
    fields = [
        "case_id",
        "source_batch_id",
        "source_bundle",
        "source_output",
        "turn_count",
        "worker_user_id",
        "primary_model",
        "primary_status",
        "review_model",
        "review_status",
        "review_item_count",
        "status",
        "elapsed_seconds",
        "case_folder",
        "case_card_path",
    ]
    with (output_dir / "batch_case_manifest.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_human_review(path: Path, rows: list[dict[str, Any]], review_rules: dict[str, Any]) -> None:
    write_human_review_index(path.with_name("human_review_index.json"), rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "human_review"
    ws.append(REVIEW_FIELDS)
    for row in rows:
        ws.append([row.get(field, "") for field in REVIEW_FIELDS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    widths = {
        "A": 14,
        "B": 72,
        "C": 24,
        "D": 22,
        "E": 72,
        "F": 70,
        "G": 45,
        "H": 45,
        "I": 45,
        "J": 45,
        "K": 18,
        "L": 55,
        "M": 40,
        "N": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if rows:
        choices = review_rules.get("human_review_choices", [])
        validation = DataValidation(type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"K2:K{len(rows) + 1}")

    guide = wb.create_sheet("how_to_fill")
    guide.append(["字段", "说明"])
    guide.append(["your_choice", "下拉选择。模型冲突时可确认 DeepSeek、确认 Qwen，或选择手工修正。"])
    guide.append(["corrected_value", "选择“手工修正”时填写最终值。可以写中文，也可以写 JSON。"])
    guide.append(["original_path", "最原始输入文件位置，比如 HTML/PDF/长图路径。"])
    guide.append(["source_images", "这一条复核项对应的第一阶段切片图路径，通常比原文件更容易定位具体位置。"])
    guide.append(["source_excerpt", "关键 turn 附近上下文，方便你不用回看整套聊天也能判断。"])
    guide.append(["notes", "你的补充说明。"])
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_human_review(path: Path, rows: list[dict[str, Any]], review_rules: dict[str, Any]) -> None:
    write_human_review_index(path.with_name("human_review_index.json"), rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "human_review"
    ws.append(REVIEW_FIELDS)
    for row in rows:
        ws.append([row.get(field, "") for field in REVIEW_FIELDS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    field_widths = {
        "review_id": 14,
        "review_priority": 18,
        "review_scope": 70,
        "review_question": 70,
        "suggested_format": 38,
        "original_path": 60,
        "case_id": 36,
        "field_cn": 24,
        "turn_ids": 22,
        "source_images": 64,
        "source_excerpt": 70,
        "deepseek_value": 45,
        "qwen_value": 45,
        "why_uncertain": 45,
        "impact_if_wrong": 45,
        "your_choice": 18,
        "corrected_value": 55,
        "notes": 40,
        "status": 12,
    }
    for index, field in enumerate(REVIEW_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = field_widths.get(field, 24)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if rows:
        choices = review_rules.get("human_review_choices", [])
        validation = DataValidation(type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True)
        ws.add_data_validation(validation)
        choice_col = get_column_letter(REVIEW_FIELDS.index("your_choice") + 1)
        validation.add(f"{choice_col}2:{choice_col}{len(rows) + 1}")

    guide = wb.create_sheet("how_to_fill")
    guide.append(["字段", "说明"])
    guide_rows = [
        ("review_priority", "复核优先级。must_review 是建议优先看的；optional_review 可以晚点看或跳过。"),
        ("review_scope", "判断范围。会说明这一行是判断某组阶段证据、某个字段，还是某个具体 turn。阶段类不是让你给整套聊天只定一个阶段。"),
        ("review_question", "这一行具体要你判断什么。阶段类会写出当前主阶段、策略阶段和阶段范围，但你判断的是当前证据片段/策略判断点。"),
        ("suggested_format", "建议填写格式。例如阶段范围写 [3,4]；gold 可以写 turn_id：原句。"),
        ("original_path", "最原始输入文件位置，例如 HTML/PDF/长图路径。"),
        ("case_id", "案例编号。用于和 cases/{case_id}/case_card.json 对应。"),
        ("field_cn", "这条复核项涉及的字段中文名。"),
        ("turn_ids", "模型判断引用的关键 turn_id。"),
        ("source_images", "对应的第一阶段切片图路径，通常比原文件更容易定位具体位置。"),
        ("source_excerpt", "关键 turn 附近上下文。优先看这里；不够再打开 source_images 或原文件。"),
        ("deepseek_value", "DeepSeek 主判断或对应字段值。"),
        ("qwen_value", "Qwen 复核意见或对应字段值。"),
        ("why_uncertain", "为什么这条进入复核。"),
        ("impact_if_wrong", "如果这项错了，会影响什么。"),
        ("your_choice", "下拉选择。确认DeepSeek/确认Qwen/手工修正/标记为不确定/跳过。"),
        ("corrected_value", "你修正后的最终答案。不是改表头，也不是解释字段名。比如阶段范围填 [3,4]；gold 填 turn_id：原句；如果只是同意模型，这列可以留空。"),
        ("notes", "你的补充说明。"),
        ("status", "处理状态，可留空或保持 pending。"),
        ("阶段填写示例", "主阶段3，策略阶段4，范围[3,4]。这里指当前证据片段/策略判断点，不是整套完整聊天只能有一个阶段。"),
        ("gold填写示例", "turn_0123：我们之间没有别人。优先原案例真实有效回复，不要选纯刺激但不可迁移的句子。"),
    ]
    for guide_row in guide_rows:
        guide.append(list(guide_row))
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 100
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_human_review_index(path: Path, rows: list[dict[str, Any]]) -> None:
    index = {
        "schema_version": "human_review_index_v1",
        "note": "This sidecar preserves hidden fields removed from the user-facing workbook.",
        "rows": [
            {field: row.get(field, "") for field in INDEX_FIELDS}
            for row in rows
        ],
    }
    write_json(path, index)


def count_turns(case: dict[str, Any]) -> int:
    return sum(len(block.get("turns", [])) for block in case.get("blocks", []))


def main() -> None:
    parser = argparse.ArgumentParser(description="Build qingsheng case cards from chat-turns batch output.")
    parser.add_argument("--batch-id", default="batch_001_data1html_5_cases")
    parser.add_argument("--input-bundle")
    parser.add_argument("--output-batch-id")
    parser.add_argument("--case-id", action="append", default=[])
    parser.add_argument("--case-limit", type=int)
    parser.add_argument("--primary-model")
    parser.add_argument("--primary-thinking", choices=["enabled", "disabled"])
    parser.add_argument("--primary-reasoning-effort", choices=["high", "max"])
    args = parser.parse_args()
    result = run_batch(
        args.batch_id,
        args.input_bundle,
        set(args.case_id) if args.case_id else None,
        args.case_limit,
        args.output_batch_id,
        args.primary_model,
        args.primary_thinking,
        args.primary_reasoning_effort,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
