from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from workflow.common.io import read_json

from .config_loader import write_json


def read_review_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["human_review"]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if row.get("review_id"):
            rows.append(row)
    return rows


def read_index(index_path: Path) -> dict[str, dict[str, Any]]:
    if not index_path.exists():
        return {}
    data = read_json(index_path)
    return {str(row.get("review_id", "")): row for row in data.get("rows", [])}


def candidate_key(row: dict[str, Any], index_row: dict[str, Any]) -> tuple[str, str]:
    review_type = str(index_row.get("review_type") or "").strip()
    field_path = str(index_row.get("field_path") or row.get("field_cn") or "").strip()
    if review_type == "stage7_evidence_check":
        return ("stage", "阶段7误判：亲密称呼/调情不等于确立关系")
    if review_type == "overwide_stage_range" or "stage_range" in field_path:
        return ("stage", "阶段范围过宽：stage_range需要收窄")
    if "gold_reference" in field_path:
        return ("gold_reference", "gold回复选择：优先真实且可迁移的男方回复")
    if "male_profile" in field_path:
        return ("male_profile", "男生画像：避免定死人格，保留案例内倾向")
    return ("general", f"{review_type or 'manual_feedback'}：{field_path or row.get('field_cn', '')}")


def extract_candidates(xlsx_path: Path, output_path: Path | None = None) -> dict[str, Any]:
    rows = read_review_rows(xlsx_path)
    index = read_index(xlsx_path.with_name("human_review_index.json"))
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        choice = str(row.get("your_choice") or "").strip()
        corrected = str(row.get("corrected_value") or "").strip()
        notes = str(row.get("notes") or "").strip()
        if not any([choice, corrected, notes]):
            continue
        index_row = index.get(str(row.get("review_id") or ""), {})
        key = candidate_key(row, index_row)
        item = grouped.setdefault(
            key,
            {
                "category": key[0],
                "pattern": key[1],
                "suggested_rule": "",
                "status": "pending_user_approval",
                "evidence": [],
            },
        )
        item["evidence"].append(
            {
                "review_id": row.get("review_id", ""),
                "case_id": row.get("case_id", index_row.get("case_id", "")),
                "your_choice": choice,
                "corrected_value": corrected,
                "notes": notes,
            }
        )
    for item in grouped.values():
        item["suggested_rule"] = suggest_rule(item["category"], item["pattern"])
    result = {
        "schema_version": "annotation_memory_candidates_v1",
        "source_review": str(xlsx_path),
        "candidate_count": len(grouped),
        "candidates": list(grouped.values()),
    }
    if output_path:
        write_json(output_path, result)
    return result


def suggest_rule(category: str, pattern: str) -> str:
    if category == "stage":
        return "阶段判断应以聊天当下证据为准；成功结果、亲密称呼或局部性张力不能直接抬高整体阶段。"
    if category == "gold_reference":
        return "gold_reference优先选择原案例真实发生、可迁移价值高的男方回复；避免选择纯刺激、低俗或只在局部语境成立的句子。"
    if category == "male_profile":
        return "男生画像只描述当前案例中的行为倾向，必须引用证据，避免定死人格。"
    return pattern


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract candidate annotation-memory rules from a filled review workbook.")
    parser.add_argument("review_xlsx")
    parser.add_argument("--output")
    args = parser.parse_args()
    review_path = Path(args.review_xlsx)
    output_path = Path(args.output) if args.output else review_path.with_name("annotation_memory_candidates.yaml")
    result = extract_candidates(review_path, output_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
