from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config_loader import OUTPUTS_ROOT, read_json, write_json
from .pipeline import field_path_cn, normalize_stage_judgment


def rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb["human_review"]
    headers = [cell.value for cell in ws[1]]
    sidecar = review_index(path.with_name("human_review_index.json"))
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[column_index]: value for column_index, value in enumerate(values)}
        if row.get("your_choice") or row.get("corrected_value") or row.get("notes"):
            hidden = sidecar.get(str(row.get("review_id") or ""), {})
            visible_case_id = str(row.get("case_id") or "")
            hidden_case_id = str(hidden.get("case_id") or "")
            if hidden and (not visible_case_id or not hidden_case_id or visible_case_id == hidden_case_id):
                for key, value in hidden.items():
                    row.setdefault(key, value)
            rows.append(row)
    return rows


def review_index(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    data = read_json(path)
    return {str(row.get("review_id", "")): row for row in data.get("rows", [])}


def apply_review(batch_id: str, review_path: Path) -> dict[str, Any]:
    output_dir = OUTPUTS_ROOT / batch_id
    rows = rows_from_xlsx(review_path)
    applied = 0
    touched: set[str] = set()
    for row in rows:
        case_id = str(row.get("case_id") or "")
        if not case_id:
            continue
        case_path = output_dir / "cases" / case_id / "case_card.json"
        case_card = read_json(case_path)
        choice = str(row.get("your_choice") or "")
        field_path = str(row.get("field_path") or "")
        corrected_value = str(row.get("corrected_value") or "").strip()
        if not field_path:
            field_path = infer_field_path(case_card, row)
            row["field_path"] = field_path
        if choice == "确认DeepSeek":
            mark_review(case_card, row, "accepted_primary")
        elif choice == "确认Qwen":
            apply_review_value(case_card, row)
            mark_review(case_card, row, "accepted_review")
        elif choice == "手工修正" and corrected_value:
            if field_path:
                apply_manual_value(case_card, field_path, corrected_value)
            mark_review(case_card, row, "manual_corrected")
        elif corrected_value:
            if field_path:
                apply_manual_value(case_card, field_path, corrected_value)
            mark_review(case_card, row, "manual_corrected")
        elif choice == "标记为不确定":
            mark_review(case_card, row, "kept_uncertain")
        elif choice == "跳过":
            mark_review(case_card, row, "skipped")
        else:
            continue
        normalize_stage_judgment(case_card)
        case_card.setdefault("quality", {})["need_human_review"] = unresolved_review_count(case_card) > 0
        write_json(case_path, case_card)
        write_json(output_dir / "cases" / case_id / "case_quality_report.json", case_card.get("quality", {}))
        touched.add(case_id)
        applied += 1
    refresh_manifest(output_dir)
    return {"batch_id": batch_id, "rows_with_input": len(rows), "applied": applied, "touched_cases": sorted(touched)}


def parse_value(text: str) -> Any:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return text


def apply_review_value(case_card: dict[str, Any], row: dict[str, Any]) -> None:
    field_path = str(row.get("field_path") or "")
    if not field_path:
        return
    review_value = value_by_path(case_card.get("model_judgments", {}).get("review", {}), field_path)
    if review_value not in ("", None):
        set_path(case_card, field_path, review_value)


def apply_manual_value(case_card: dict[str, Any], field_path: str, corrected_value: str) -> None:
    stage_patch = parse_stage_correction(corrected_value)
    is_stage_field = "stage" in field_path or field_path in {"qingsheng_mapping.stage_number", "qingsheng_mapping.stage_label"}
    if stage_patch and is_stage_field:
        judgment = case_card.setdefault("qingsheng_mapping", {}).setdefault("stage_judgment", {})
        if stage_patch.get("primary_stage") is not None:
            judgment["primary_stage"] = stage_patch["primary_stage"]
        if stage_patch.get("strategy_stage") is not None:
            judgment["strategy_stage"] = stage_patch["strategy_stage"]
        if stage_patch.get("stage_range"):
            judgment["stage_range"] = stage_patch["stage_range"]
        return
    if is_stage_field:
        return
    set_path(case_card, field_path, parse_value(corrected_value))


def infer_field_path(case_card: dict[str, Any], row: dict[str, Any]) -> str:
    field_cn = str(row.get("field_cn") or "")
    if not field_cn:
        return ""
    for item in case_card.get("quality", {}).get("review_items", []):
        field = str(item.get("field") or "")
        if field and field_path_cn(field) == field_cn:
            return field
    fallback = {
        "阶段范围": "qingsheng_mapping.stage_judgment.stage_range",
        "主判断阶段": "qingsheng_mapping.stage_judgment.primary_stage",
        "策略阶段": "qingsheng_mapping.stage_judgment.strategy_stage",
        "关系阶段编号": "qingsheng_mapping.stage_number",
        "参考下一句": "gold_reference.next_reply",
        "原案例真实好回复": "gold_reference.observed_good_reply",
        "参考回复来源": "gold_reference.reference_type",
    }
    return fallback.get(field_cn, "")


def parse_stage_correction(text: str) -> dict[str, Any]:
    import re

    result: dict[str, Any] = {}
    primary = re.search(r"主阶段\s*([0-7])", text)
    strategy = re.search(r"策略阶段\s*([0-7])", text)
    range_match = re.search(r"\[\s*([0-7])\s*,\s*([0-7])\s*\]", text)
    loose_range = re.search(r"([0-7])\s*(?:-|—|–|,|，|到|至|穿插)+\s*([0-7])", text)
    if primary:
        result["primary_stage"] = int(primary.group(1))
    if strategy:
        result["strategy_stage"] = int(strategy.group(1))
    if range_match:
        result["stage_range"] = [int(range_match.group(1)), int(range_match.group(2))]
    elif loose_range:
        left = int(loose_range.group(1))
        right = int(loose_range.group(2))
        result["stage_range"] = [min(left, right), max(left, right)]
    return result


def set_path(data: dict[str, Any], dotted_path: str, value: Any) -> None:
    parts = [part for part in dotted_path.split(".") if part]
    if not parts:
        return
    current: Any = data
    for part in parts[:-1]:
        if not isinstance(current, dict):
            return
        current = current.setdefault(part, {})
    if isinstance(current, dict):
        current[parts[-1]] = value


def value_by_path(data: Any, dotted_path: str) -> Any:
    current = data
    for part in [part for part in dotted_path.split(".") if part]:
        if isinstance(current, dict):
            current = current.get(part, "")
        else:
            return ""
    return current


def mark_review(case_card: dict[str, Any], row: dict[str, Any], status: str) -> None:
    quality = case_card.setdefault("quality", {})
    review_id = row.get("review_id", "")
    existing = quality.setdefault("human_review_applied", [])
    quality["human_review_applied"] = [
        item for item in existing if item.get("review_id", "") != review_id
    ]
    quality["human_review_applied"].append(
        {
            "review_id": review_id,
            "field_path": row.get("field_path", ""),
            "choice": row.get("your_choice", ""),
            "corrected_value": row.get("corrected_value", ""),
            "notes": row.get("notes", ""),
            "status": status,
        }
    )
    for item in quality.get("review_items", []):
        same_field = item.get("field", "") == row.get("field_path", "")
        same_type = row.get("review_type") and item.get("type", "") == row.get("review_type", "")
        same_cn = field_path_cn(str(item.get("field", ""))) == str(row.get("field_cn") or "")
        if same_field or same_type or same_cn:
            item["human_review_status"] = status


def unresolved_review_count(case_card: dict[str, Any]) -> int:
    count = 0
    for item in case_card.get("quality", {}).get("review_items", []):
        if not item.get("human_review_status"):
            count += 1
    return count


def refresh_manifest(output_dir: Path) -> None:
    manifest_path = output_dir / "batch_case_manifest.json"
    if not manifest_path.exists():
        return
    manifest = read_json(manifest_path)
    for row in manifest.get("cases", []):
        case_path = output_dir / "cases" / row["case_id"] / "case_card.json"
        if not case_path.exists():
            continue
        case_card = read_json(case_path)
        review_count = unresolved_review_count(case_card)
        row["review_item_count"] = review_count
        row["status"] = "needs_human_review" if review_count else "ready"
    write_json(manifest_path, manifest)


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply human review choices to qingsheng case cards.")
    parser.add_argument("--batch-id", default="batch_001_data1html_5_cases")
    parser.add_argument("--review-xlsx")
    args = parser.parse_args()
    review_path = Path(args.review_xlsx) if args.review_xlsx else OUTPUTS_ROOT / args.batch_id / "human_review.xlsx"
    print(json.dumps(apply_review(args.batch_id, review_path), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
