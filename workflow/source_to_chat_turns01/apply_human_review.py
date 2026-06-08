from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CHOICE_TO_SPEAKER = {
    "改为男生": "male",
    "改为女生": "female",
    "男生表情包": "male",
    "女生表情包": "female",
    "女生自拍照片": "female",
    "男生生活照": "male",
    "普通图片": "unknown",
    "复盘/讲解": "narration",
    "改为旁白": "narration",
    "改为系统": "system",
    "无法判断": "unknown",
}

STICKER_CHOICES = {"男生表情包", "女生表情包"}
CHOICE_TO_CONTENT_TYPE = {
    "男生表情包": "sticker",
    "女生表情包": "sticker",
    "女生自拍照片": "selfie_photo",
    "男生生活照": "life_photo",
    "普通图片": "image",
    "复盘/讲解": "narration",
    "改为旁白": "narration",
    "改为系统": "system",
}
CHOICE_TO_VISUAL_NOTE = {
    "男生表情包": "男生发送的表情包",
    "女生表情包": "女生发送的表情包",
    "女生自拍照片": "女生发送的自拍/个人照片",
    "男生生活照": "男生发送的生活照",
    "普通图片": "聊天中的普通图片",
    "复盘/讲解": "复盘/讲解内容",
}

EMPTY_TEXT_MARKERS = {"空白无内容", "空白 无内容", "无内容", "没有内容", "空白", "跳过"}
SPEAKER_PREFIX = {
    "男": "male",
    "男生": "male",
    "女": "female",
    "女生": "female",
    "旁白": "narration",
    "系统": "system",
}


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def compact_marker(text: str) -> str:
    return re.sub(r"\s+", "", text.strip())


def is_empty_marker(text: str) -> bool:
    return compact_marker(text) in {compact_marker(item) for item in EMPTY_TEXT_MARKERS}


def parse_manual_turns(text: str, default_speaker: str) -> list[dict[str, str]]:
    turns: list[dict[str, str]] = []
    current_time = ""
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if re.fullmatch(r"\d{4}年\d{1,2}月\d{1,2}日\s*(上午|下午|晚上|中午)?\d{1,2}:\d{2}", line):
            current_time = line
            continue
        match = re.match(r"^(男生|女生|男|女|旁白|系统)\s*[:：]\s*(.*)$", line)
        if match:
            turns.append(
                {
                    "speaker": SPEAKER_PREFIX.get(match.group(1), default_speaker or "unknown"),
                    "text": match.group(2).strip(),
                    "time": current_time,
                }
            )
            continue
        if turns:
            turns[-1]["text"] = f"{turns[-1]['text']}\n{line}".strip()
        else:
            turns.append({"speaker": default_speaker or "unknown", "text": line, "time": current_time})
    return [turn for turn in turns if turn["text"]]


def rows_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb["human_review"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values)}
        if row.get("your_choice") or row.get("corrected_speaker") or row.get("corrected_text"):
            rows.append(row)
    return rows


def next_turn_id(blocks: list[dict[str, Any]]) -> str:
    max_id = 0
    for block in blocks:
        for turn in block.get("turns", []):
            value = str(turn.get("turn_id", "")).removeprefix("turn_")
            if value.isdigit():
                max_id = max(max_id, int(value))
    return f"turn_{max_id + 1:04d}"


def prepared_block(batch_dir: Path, case_id: str, block_id: str) -> dict[str, Any]:
    manifest_path = batch_dir / "cases" / case_id / "block_manifest.json"
    if not manifest_path.exists():
        return {"block_id": block_id, "turns": []}
    manifest = read_json(manifest_path)
    for block in manifest.get("blocks", []):
        if block.get("block_id") == block_id:
            return {
                "block_id": block_id,
                "image_quality": "manual_review",
                "extracted_text": "",
                "turns": [],
                "needs_human_review": False,
                "issues": ["model call blocked; manually reviewed"],
                "source_image": block.get("prepared_path", ""),
                "crop_box": block.get("crop_box", []),
            }
    return {"block_id": block_id, "turns": []}


def ensure_block(blocks: list[dict[str, Any]], batch_dir: Path, case_id: str, block_id: str) -> dict[str, Any]:
    for block in blocks:
        if block.get("block_id") == block_id:
            return block
    block = prepared_block(batch_dir, case_id, block_id)
    blocks.append(block)
    blocks.sort(key=lambda item: int(str(item.get("block_id", "block_999999")).split("_")[-1]))
    return block


def apply_rows(batch_dir: Path, review_path: Path) -> dict[str, Any]:
    rows = rows_from_xlsx(review_path)
    touched: set[str] = set()
    handled_failed_blocks: dict[str, set[str]] = {}
    applied = 0
    for row in rows:
        case_id = str(row.get("case_id") or "")
        if not case_id:
            continue
        case_path = batch_dir / "cases" / case_id / "chat_turns.json"
        data = read_json(case_path)
        blocks = data.get("blocks", [])
        review_type = str(row.get("review_type") or "")
        block_id = str(row.get("block_id") or "")
        turn_id = str(row.get("turn_id") or "")
        choice = str(row.get("your_choice") or "")
        corrected_speaker = str(row.get("corrected_speaker") or "").strip()
        corrected_text = str(row.get("corrected_text") or "").strip()

        if review_type in {"unknown_turn", "narration_suspect"}:
            target_speaker = corrected_speaker or CHOICE_TO_SPEAKER.get(choice, "")
            for block in blocks:
                if block.get("block_id") != block_id:
                    continue
                turns = block.get("turns", [])
                for turn in list(turns):
                    if turn.get("turn_id") == turn_id:
                        if choice == "跳过":
                            turns.remove(turn)
                            applied += 1
                            touched.add(case_id)
                            break
                        if target_speaker:
                            turn["speaker"] = target_speaker
                        content_type = CHOICE_TO_CONTENT_TYPE.get(choice, "")
                        if content_type:
                            turn["content_type"] = content_type
                        visual_note = CHOICE_TO_VISUAL_NOTE.get(choice, "")
                        if visual_note:
                            turn["visual_note"] = visual_note
                        if choice in STICKER_CHOICES and not corrected_text:
                            turn["text"] = "[表情包]"
                        elif choice == "女生自拍照片" and not corrected_text:
                            turn["text"] = "[女生自拍照片]"
                        elif choice == "男生生活照" and not corrected_text:
                            turn["text"] = "[男生生活照]"
                        elif choice == "普通图片" and not corrected_text:
                            turn["text"] = "[图片]"
                        if corrected_text:
                            turn["text"] = corrected_text
                        turn["need_review"] = False
                        turn["notes"] = "; ".join(
                            part
                            for part in [
                                str(turn.get("notes", "")).strip(),
                                f"human_review: {review_type} corrected",
                                f"human_review: {content_type}" if content_type else "",
                            ]
                            if part
                        )
                        applied += 1
                        touched.add(case_id)

        if review_type == "model_blocked_block" and corrected_text:
            block = ensure_block(blocks, batch_dir, case_id, block_id)
            if not is_empty_marker(corrected_text):
                default_speaker = corrected_speaker or CHOICE_TO_SPEAKER.get(choice, "unknown")
                if choice in STICKER_CHOICES and is_empty_marker(corrected_text):
                    corrected_text = "男生：[表情包]" if default_speaker == "male" else "女生：[表情包]"
                for manual_turn in parse_manual_turns(corrected_text, default_speaker):
                    block.setdefault("turns", []).append(
                        {
                            "turn_id": next_turn_id(blocks),
                            "speaker": manual_turn["speaker"],
                            "text": manual_turn["text"],
                            "time": manual_turn.get("time", ""),
                            "confidence": "human",
                            "reason": "manual transcript for model-blocked image",
                            "source_block_id": block_id,
                            "source_image": block.get("source_image", ""),
                            "crop_box": block.get("crop_box", []),
                            "need_review": False,
                            "content_type": CHOICE_TO_CONTENT_TYPE.get(choice, "text"),
                            "visual_note": CHOICE_TO_VISUAL_NOTE.get(choice, ""),
                            "notes": "human_review_applied; original model call blocked by provider safety policy"
                            + (f"; human_review: {CHOICE_TO_CONTENT_TYPE.get(choice)}" if CHOICE_TO_CONTENT_TYPE.get(choice) else ""),
                        }
                    )
                block["extracted_text"] = corrected_text
            else:
                block["extracted_text"] = ""
                block["issues"] = ["human confirmed blank/no content"]
            block["needs_human_review"] = False
            handled_failed_blocks.setdefault(case_id, set()).add(block_id)
            applied += 1
            touched.add(case_id)

        write_json(case_path, data)

    update_manifest(batch_dir, handled_failed_blocks)
    for case_id in touched:
        case_path = batch_dir / "cases" / case_id / "chat_turns.json"
        if case_path.exists():
            data = read_json(case_path)
            remove_adjacent_duplicate_turns(data.get("blocks", []))
            write_json(case_path, data)
    rebuild_batch(batch_dir)
    return {"rows_with_input": len(rows), "applied": applied, "touched_cases": sorted(touched)}


def normalize_turn_text(text: str) -> str:
    return re.sub(r"\s+", "", text)


def remove_adjacent_duplicate_turns(blocks: list[dict[str, Any]]) -> None:
    previous_key = None
    previous_human = False
    for block in blocks:
        kept = []
        for turn in block.get("turns", []):
            key = (turn.get("speaker"), normalize_turn_text(str(turn.get("text", ""))))
            is_human = turn.get("confidence") == "human"
            if key[1] and key == previous_key and is_human:
                continue
            kept.append(turn)
            previous_key = key
            previous_human = is_human
        block["turns"] = kept


def update_manifest(batch_dir: Path, handled_failed_blocks: dict[str, set[str]]) -> None:
    manifest_path = batch_dir / "batch_manifest.json"
    manifest = read_json(manifest_path)
    for case in manifest.get("cases", []):
        handled = handled_failed_blocks.get(case.get("case_id", ""), set())
        if handled:
            remaining_failures = []
            for failure in case.get("failure_blocks", []):
                remaining_block_ids = [block_id for block_id in failure.get("block_ids", []) if block_id not in handled]
                if remaining_block_ids:
                    updated = dict(failure)
                    updated["block_ids"] = remaining_block_ids
                    remaining_failures.append(updated)
            case["failure_blocks"] = remaining_failures
            case["failed_group_count"] = len(remaining_failures)
            case["failure_count"] = len(remaining_failures)
        case_path = batch_dir / "cases" / case["case_id"] / "chat_turns.json"
        if case_path.exists():
            quality = refresh_case_quality(batch_dir, case)
            case["need_review_turns"] = quality.get("need_review_turns", 0)
            case["speaker_counts"] = quality.get("speaker_counts", {})
        if case.get("failed_group_count", case.get("failure_count", 0)) == 0 and case.get("need_review_turns", 0) == 0:
            case["status"] = "ready"
        elif str(case.get("status", "")).startswith("deferred"):
            pass
        else:
            case["status"] = "needs_attention"
    write_json(manifest_path, manifest)
    write_manifest_csv(batch_dir, manifest)


def refresh_case_quality(batch_dir: Path, case: dict[str, Any]) -> dict[str, Any]:
    case_id = case["case_id"]
    case_dir = batch_dir / "cases" / case_id
    chat_path = case_dir / "chat_turns.json"
    quality_path = case_dir / "quality_report.json"
    raw_path = case_dir / "raw_model_results.json"
    data = read_json(chat_path)
    previous = read_json(quality_path) if quality_path.exists() else data.get("summary", {})
    raw = read_json(raw_path) if raw_path.exists() else {"results": []}
    blocks = data.get("blocks", [])
    ensure_content_fields(blocks)
    results = raw.get("results", [])
    quality = {
        **previous,
        "case_id": case_id,
        "image_count": previous.get("image_count", 0),
        "call_count": previous.get("call_count", len(results)),
        "success_count": sum(1 for item in results if item.get("status") == "model_success")
        if results
        else previous.get("success_count", 0),
        "failure_count": sum(1 for item in results if item.get("status") != "model_success")
        if results
        else previous.get("failure_count", 0),
        "status_counts": status_counts(results) if results else previous.get("status_counts", {}),
        "speaker_counts": speaker_counts(blocks),
        "content_type_counts": content_type_counts(blocks),
        "need_review_turns": sum(
            1 for block in blocks for turn in block.get("turns", []) if turn.get("need_review")
        ),
    }
    data["summary"] = quality
    write_json(chat_path, data)
    write_json(quality_path, quality)
    return quality


def ensure_content_fields(blocks: list[dict[str, Any]]) -> None:
    for block in blocks:
        for turn in block.get("turns", []):
            if turn.get("content_type"):
                if (
                    turn.get("content_type") == "image"
                    and "human_review: image" not in str(turn.get("notes", ""))
                    and not str(turn.get("text", "")).strip().startswith(("[图片]", "[模糊头像图片]"))
                    and "图片" not in str(turn.get("reason", ""))
                    and "照片" not in str(turn.get("reason", ""))
                ):
                    turn["content_type"] = "text"
                    turn["visual_note"] = ""
                turn.setdefault("visual_note", "")
                continue
            speaker = str(turn.get("speaker", "unknown"))
            text = str(turn.get("text", "")).strip()
            reason = str(turn.get("reason", "")).strip()
            if speaker == "system":
                turn["content_type"] = "system"
                turn.setdefault("visual_note", "系统提示")
            elif speaker == "narration":
                turn["content_type"] = "narration" if "[表情包]" not in text and "[图片]" not in text else "image"
                turn.setdefault("visual_note", reason or "复盘/讲解内容")
            elif text in {"[表情包]", "[贴纸]"} or "表情包" in reason or "贴纸" in reason or "动态图" in reason:
                turn["content_type"] = "sticker"
                turn.setdefault("visual_note", "聊天气泡中的表情包/贴纸")
            elif (text.startswith("[女生自拍照片]") or "自拍" in reason or "个人照片" in reason) and speaker == "female":
                turn["content_type"] = "selfie_photo"
                turn.setdefault("visual_note", "女生发送的自拍/个人照片")
            elif (text.startswith("[男生生活照]") or "生活照" in reason) and speaker == "male":
                turn["content_type"] = "life_photo"
                turn.setdefault("visual_note", "男生发送的生活照")
            elif text.startswith("[图片]") or text.startswith("[模糊头像图片]") or "图片" in reason or "照片" in reason:
                turn["content_type"] = "image"
                turn.setdefault("visual_note", "聊天中的图片")
            elif text:
                turn["content_type"] = "text"
                turn.setdefault("visual_note", "")
            else:
                turn["content_type"] = "unknown"
                turn.setdefault("visual_note", "")


def speaker_counts(blocks: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for block in blocks:
        for turn in block.get("turns", []):
            speaker = str(turn.get("speaker", "unknown"))
            counts[speaker] = counts.get(speaker, 0) + 1
    return counts


def content_type_counts(blocks: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for block in blocks:
        for turn in block.get("turns", []):
            content_type = str(turn.get("content_type", "unknown"))
            counts[content_type] = counts.get(content_type, 0) + 1
    return counts


def status_counts(results: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in results:
        status = str(item.get("status", "unknown"))
        counts[status] = counts.get(status, 0) + 1
    return counts


def write_manifest_csv(batch_dir: Path, manifest: dict[str, Any]) -> None:
    fields = [
        "case_id",
        "source_output",
        "mode",
        "image_count",
        "call_count",
        "success_count",
        "failure_count",
        "failed_group_count",
        "need_review_turns",
        "status",
        "case_folder",
        "chat_turns_path",
        "quality_report_path",
    ]
    with (batch_dir / "batch_manifest.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for case in manifest.get("cases", []):
            writer.writerow({field: case.get(field, "") for field in fields})


def rebuild_batch(batch_dir: Path) -> None:
    manifest = read_json(batch_dir / "batch_manifest.json")
    combined = []
    for case in manifest.get("cases", []):
        case_id = case["case_id"]
        case_path = batch_dir / "cases" / case_id / "chat_turns.json"
        turns = read_json(case_path)
        combined.append(
            {
                "case_id": case_id,
                "summary": turns.get("summary", {}),
                "blocks": turns.get("blocks", []),
                "failure_blocks": case.get("failure_blocks", []),
            }
        )
    write_json(
        batch_dir / "batch_chat_turns.json",
        {
            "schema_version": "chat_turns_batch_v1",
            "batch_id": manifest.get("batch_id", batch_dir.name),
            "cases": combined,
        },
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply human review workbook edits to a chat-turns batch.")
    parser.add_argument("batch_dir")
    parser.add_argument("review_xlsx")
    args = parser.parse_args()
    result = apply_rows(Path(args.batch_dir), Path(args.review_xlsx))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
