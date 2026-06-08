from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


FIELDS = [
    "review_id",
    "review_type",
    "case_id",
    "block_id",
    "turn_id",
    "source_image",
    "current_speaker",
    "current_text",
    "context_before",
    "context_after",
    "block_extracted_text",
    "model_reason",
    "question_cn",
    "your_choice",
    "corrected_speaker",
    "corrected_text",
    "notes",
    "status",
]

CHOICES = [
    "确认AI判断",
    "改为男生",
    "改为女生",
    "男生表情包",
    "女生表情包",
    "女生自拍照片",
    "男生生活照",
    "普通图片",
    "复盘/讲解",
    "改为旁白",
    "改为系统",
    "手工转写",
    "跳过",
    "无法判断",
]

SPEAKER_MAP = {
    "改为男生": "male",
    "改为女生": "female",
    "男生表情包": "male",
    "女生表情包": "female",
    "女生自拍照片": "female",
    "男生生活照": "male",
    "普通图片": "unknown",
    "复盘/讲解": "narration",
    "改为旁白": "narration",
    "改为系统": "system",
}

NARRATION_TRIGGERS = [
    "我和大家说一下",
    "兄弟们",
    "这个案例",
    "开撩",
    "复盘",
    "总结",
    "这里要注意",
    "这一段",
    "这个女生",
    "这个男生",
]


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def source_image_for_block(batch_dir: Path, case_dir: Path, case_id: str, block_id: str) -> str:
    data = read_json(case_dir / "chat_turns.json")
    for block in data.get("blocks", []):
        if block.get("block_id") == block_id:
            if block.get("source_image"):
                return str(block["source_image"])
            for turn in block.get("turns", []):
                if turn.get("source_image"):
                    return str(turn["source_image"])
    prepared_manifest = batch_dir / "cases" / case_id / "block_manifest.json"
    if prepared_manifest.exists():
        data = read_json(prepared_manifest)
        for block in data.get("blocks", []):
            if block.get("block_id") == block_id:
                return str(block.get("prepared_path", ""))
    return ""


def build_rows(batch_dir: Path) -> list[dict[str, Any]]:
    manifest = read_json(batch_dir / "batch_manifest.json")
    rows: list[dict[str, Any]] = []
    index = 1
    for case in manifest.get("cases", []):
        case_id = case["case_id"]
        case_dir = batch_dir / "cases" / case_id
        turns_data = read_json(case_dir / "chat_turns.json")
        for block in turns_data.get("blocks", []):
            turns = block.get("turns", [])
            for turn_index, turn in enumerate(turns):
                if turn.get("need_review") or turn.get("speaker") == "unknown":
                    before = turns[max(0, turn_index - 3) : turn_index]
                    after = turns[turn_index + 1 : turn_index + 4]
                    rows.append(
                        {
                            "review_id": f"review_{index:04d}",
                            "review_type": "unknown_turn",
                            "case_id": case_id,
                            "block_id": block.get("block_id", ""),
                            "turn_id": turn.get("turn_id", ""),
                            "source_image": turn.get("source_image", block.get("source_image", "")),
                            "current_speaker": turn.get("speaker", ""),
                            "current_text": turn.get("text", ""),
                            "context_before": "\n".join(
                                f"{item.get('speaker', '')}：{item.get('text', '')}" for item in before
                            ),
                            "context_after": "\n".join(
                                f"{item.get('speaker', '')}：{item.get('text', '')}" for item in after
                            ),
                            "block_extracted_text": str(block.get("extracted_text", "")),
                            "model_reason": str(turn.get("reason", "")),
                            "question_cn": "这句话/区域说话人或内容类型不确定。请看 context_before/context_after 和 source_image 定位；如果是表情包、女生自拍照、男生生活照、普通图片或复盘/讲解，直接选对应项；如果 current_text 为空，通常是空白气泡/头像/误识别，可选“跳过”；如果确实有文字，请在 corrected_text 写正确文字。",
                            "your_choice": "",
                            "corrected_speaker": "",
                            "corrected_text": "",
                            "notes": "",
                            "status": "pending",
                        }
                    )
                    index += 1
                elif is_narration_suspect(turn):
                    before = turns[max(0, turn_index - 3) : turn_index]
                    after = turns[turn_index + 1 : turn_index + 4]
                    suggested = suggested_speaker(turn)
                    rows.append(
                        {
                            "review_id": f"review_{index:04d}",
                            "review_type": "narration_suspect",
                            "case_id": case_id,
                            "block_id": block.get("block_id", ""),
                            "turn_id": turn.get("turn_id", ""),
                            "source_image": turn.get("source_image", block.get("source_image", "")),
                            "current_speaker": turn.get("speaker", ""),
                            "current_text": turn.get("text", ""),
                            "context_before": "\n".join(
                                f"{item.get('speaker', '')}：{item.get('text', '')}" for item in before
                            ),
                            "context_after": "\n".join(
                                f"{item.get('speaker', '')}：{item.get('text', '')}" for item in after
                            ),
                            "block_extracted_text": str(block.get("extracted_text", "")),
                            "model_reason": str(turn.get("reason", "")),
                            "question_cn": f"模型标为旁白，但不像讲解/复盘文本，疑似错标（建议改为：{speaker_label(suggested)}）。如果这是表情包、女生自拍照、男生生活照或普通图片，请直接选对应项。请根据上下文确认。",
                            "your_choice": "",
                            "corrected_speaker": "",
                            "corrected_text": "",
                            "notes": "",
                            "status": "pending",
                        }
                    )
                    index += 1
        for failure in case.get("failure_blocks", []):
            for block_id in failure.get("block_ids", []):
                rows.append(
                    {
                        "review_id": f"review_{index:04d}",
                        "review_type": "model_blocked_block",
                        "case_id": case_id,
                        "block_id": block_id,
                        "turn_id": "",
                        "source_image": source_image_for_block(batch_dir, case_dir, case_id, block_id),
                        "current_speaker": "",
                        "current_text": "",
                        "context_before": "",
                        "context_after": "",
                        "block_extracted_text": "",
                        "model_reason": "",
                        "question_cn": "这张图片被模型安全策略拦截，没有生成话轮。请打开 source_image 人工转写；如果是表情包/照片/普通图片，可直接选对应项；如果有文字，一行可写多句，建议格式：男生：... / 女生：... / 旁白：...",
                        "your_choice": "",
                        "corrected_speaker": "",
                        "corrected_text": "",
                        "notes": str(failure.get("error", ""))[:500],
                        "status": "pending",
                    }
                )
                index += 1
    return rows


def is_narration_suspect(turn: dict[str, Any]) -> bool:
    if turn.get("speaker") != "narration":
        return False
    notes = str(turn.get("notes", ""))
    if "human_review" in notes:
        return False
    text = str(turn.get("text", "")).strip()
    if any(trigger in text for trigger in NARRATION_TRIGGERS):
        return False
    reason = str(turn.get("reason", ""))
    return any(marker in reason for marker in ["左侧", "右侧", "白色", "绿色", "气泡"]) or bool(text)


def suggested_speaker(turn: dict[str, Any]) -> str:
    reason = str(turn.get("reason", ""))
    text = str(turn.get("text", "")).strip()
    if any(marker in reason for marker in ["右侧", "绿色", "本人"]):
        return "male"
    if any(marker in reason for marker in ["左侧", "白色", "对方"]):
        return "female"
    if text.startswith("[图片]") or text.startswith("[音频]") or text.startswith("0:"):
        return "system"
    return "female"


def speaker_label(speaker: str) -> str:
    return {
        "male": "男生",
        "female": "女生",
        "narration": "旁白",
        "system": "系统",
        "unknown": "不确定",
    }.get(speaker, "不确定")


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "human_review"
    ws.append(FIELDS)
    for row in rows:
        ws.append([row.get(field, "") for field in FIELDS])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 14,
        "B": 22,
        "C": 34,
        "D": 14,
        "E": 14,
        "F": 72,
        "G": 16,
        "H": 36,
        "I": 42,
        "J": 42,
        "K": 60,
        "L": 36,
        "M": 62,
        "N": 18,
        "O": 18,
        "P": 60,
        "Q": 44,
        "R": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if rows:
        choice_validation = DataValidation(type="list", formula1='"' + ",".join(CHOICES) + '"', allow_blank=True)
        ws.add_data_validation(choice_validation)
        choice_validation.add(f"N2:N{len(rows) + 1}")
        speaker_validation = DataValidation(
            type="list",
            formula1='"male,female,narration,system,unknown"',
            allow_blank=True,
        )
        ws.add_data_validation(speaker_validation)
        speaker_validation.add(f"O2:O{len(rows) + 1}")

    guide = wb.create_sheet("how_to_fill")
    guide.append(["字段", "说明"])
    guide_rows = [
        ("your_choice", "下拉选择处理方式。普通文字选“改为男生/女生/旁白/系统”；图片类选“男生表情包/女生表情包/女生自拍照片/男生生活照/普通图片”；复盘讲解选“复盘/讲解”；无法识别内容可选“跳过/无法判断”。"),
        ("corrected_speaker", "可选。下拉选择 male/female/narration/system/unknown；会优先于 your_choice。"),
        ("corrected_text", "可选。unknown_turn 用于改错字；model_blocked_block 必填，用于人工转写整张图。"),
        ("source_image", "本地图片路径。需要时复制这个路径打开原图定位。"),
        ("notes", "可以写补充说明；安全拦截行这里会带模型失败原因。"),
    ]
    for item in guide_rows:
        guide.append(list(item))
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a combined human review workbook for a chat-turns batch.")
    parser.add_argument("batch_dir")
    parser.add_argument("--output")
    args = parser.parse_args()
    batch_dir = Path(args.batch_dir)
    output = Path(args.output) if args.output else batch_dir / "batch_001_human_review.xlsx"
    rows = build_rows(batch_dir)
    write_xlsx(output, rows)
    print(json.dumps({"output": str(output), "rows": len(rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
