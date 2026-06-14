from pathlib import Path

from openpyxl import load_workbook

from baiou.case_pipeline.production.apply_review import apply_manual_review_fields, rows_from_workbook
from baiou.case_pipeline.production.build_segments import apply_review_policy, rows_for_review, write_review_workbook


def test_review_rows_include_v02_label_columns() -> None:
    segment = {
        "case_id": "case_001",
        "segment_id": "case_001_seg_001",
        "source_turn_ids": ["turn_0001"],
        "当前上下文": "双方已经见过面，但聊天仍在熟悉。",
        "女生最后一句": "那下次再说",
        "男生原回复": "行，下次当面聊",
        "原回复评价": "有效：低压力推进。",
        "聊天阶段": "熟悉期",
        "接触状态": "已线下接触",
        "关系推进目标": "约会后跟进",
        "女生状态": "正常",
        "男生目标": "邀约",
        "推荐策略": "模糊邀约",
        "风险类型": [],
        "回复强度": "推进",
        "高热度信号": "无",
        "更优回复": "保留原回复：行，下次当面聊",
        "迁移学习价值": "见过面后仍可低压力承接。",
        "need_human_review": True,
        "model_review": {
            "verdict": "revise",
            "need_human_review": True,
            "issues": [
                {"field": "接触状态", "current_value": "已线下接触", "suggested_value": "多次线下接触", "reason": "上下文显示多次见面。"},
                {"field": "关系推进目标", "current_value": "约会后跟进", "suggested_value": "接触后关系升级", "reason": "本段重点是升级关系。"},
            ],
        },
    }

    row = rows_for_review("case_001", [segment], {}, 1, {})[0]

    assert row["主模型接触状态"] == "已线下接触"
    assert row["主模型关系推进目标"] == "约会后跟进"
    assert row["复核建议接触状态"] == "多次线下接触"
    assert row["复核建议关系推进目标"] == "接触后关系升级"
    assert "接触状态：已线下接触" in row["主模型标签"]


def test_manual_v02_label_columns_are_applied() -> None:
    segment = {"接触状态": "未知", "关系推进目标": "无"}
    row = {"人工接触状态": "已线下接触", "人工关系推进目标": "性关系推进"}

    assert apply_manual_review_fields(segment, row)
    assert segment["接触状态"] == "已线下接触"
    assert segment["关系推进目标"] == "性关系推进"


def test_optional_review_sheet_separates_low_impact_issues(tmp_path: Path) -> None:
    must_segment = {
        "case_id": "case_001",
        "segment_id": "case_001_seg_001",
        "source_turn_ids": ["turn_0001"],
        "当前上下文": "女生拒绝后男生继续强推。",
        "女生最后一句": "不想去",
        "男生原回复": "你必须来",
        "原回复评价": "有问题：强行邀约。",
        "聊天阶段": "熟悉期",
        "接触状态": "未线下见面",
        "关系推进目标": "邀约见面",
        "女生状态": "拒绝",
        "男生目标": "邀约",
        "推荐策略": "明确邀约",
        "风险类型": [],
        "回复强度": "推进",
        "高热度信号": "无",
        "更优回复": "那先不急，下次舒服点再说",
        "迁移学习价值": "拒绝后要降压。",
        "need_human_review": True,
        "model_review": {
            "verdict": "revise",
            "need_human_review": True,
            "issues": [{"field": "风险类型", "current_value": [], "suggested_value": ["强行邀约"], "reason": "女生已经拒绝。"}],
        },
    }
    optional_segment = {
        **must_segment,
        "segment_id": "case_001_seg_002",
        "need_human_review": False,
        "model_review": {
            "verdict": "revise",
            "need_human_review": False,
            "issues": [{"field": "次要标签.说明", "current_value": "", "suggested_value": "可补充边界说明", "reason": "只是辅助说明。"}],
        },
    }
    rows = rows_for_review("case_001", [must_segment, optional_segment], {}, 1, {})
    workbook_path = tmp_path / "review.xlsx"
    write_review_workbook(workbook_path, rows, [])

    wb = load_workbook(workbook_path)
    assert wb["segments_review"].max_row == 2
    assert wb["optional_review"].max_row == 2
    assert wb["segments_review"]["T2"].value == "必须复核"
    assert wb["optional_review"]["T2"].value == "可选复核"

    rows = rows_from_workbook(workbook_path)
    assert {row["segment_id"] for row in rows} == {"case_001_seg_001", "case_001_seg_002"}


def test_optional_review_only_segment_is_auto_approved() -> None:
    segments = [
        {
            "更优回复": "可以这样说",
            "quality_status": "draft",
            "need_human_review": True,
            "model_review": {
                "verdict": "revise",
                "issues": [{"field": "次要标签.说明", "suggested_value": "边界说明", "reason": "辅助判断。"}],
            },
        }
    ]

    assert apply_review_policy(segments)
    assert segments[0]["quality_status"] == "approved"
    assert segments[0]["need_human_review"] is False
    assert "可选复核表" in segments[0]["auto_review_reason"]


def test_effective_segment_without_review_issues_is_auto_approved() -> None:
    segments = [
        {
            "原回复评价": "有效：自然接住，低压力延续。",
            "更优回复": "可以这样说",
            "quality_status": "draft",
            "need_human_review": True,
            "model_review": {},
        }
    ]

    assert apply_review_policy(segments)
    assert segments[0]["quality_status"] == "approved"
    assert segments[0]["need_human_review"] is False
    assert "原回复评价为有效" in segments[0]["auto_review_reason"]


def test_review_workbook_sorts_high_priority_rows_before_label_only(tmp_path: Path) -> None:
    label_row = {
        "review_id": "review_0001",
        "case_id": "case_001",
        "segment_id": "seg_label",
        "主模型原回复评价": "有效：回复可保留。",
        "复核模型结论": "复核结论：建议修改",
        "复核模型修改建议": "1. 字段：聊天阶段\n   复核建议值：熟悉期",
        "复核分层": "必须复核",
        "分层理由": "涉及核心字段：聊天阶段",
    }
    risk_row = {
        "review_id": "review_0002",
        "case_id": "case_001",
        "segment_id": "seg_risk",
        "主模型原回复评价": "失败：强行推进。",
        "复核模型结论": "复核结论：建议修改",
        "复核模型修改建议": "1. 字段：风险类型\n   复核建议值：[\"强行暧昧\"]",
        "复核分层": "必须复核",
        "分层理由": "涉及核心字段：风险类型",
    }
    workbook_path = tmp_path / "review.xlsx"

    write_review_workbook(workbook_path, [label_row, risk_row], [])

    wb = load_workbook(workbook_path)
    assert wb["segments_review"]["A2"].value == "review_0002"
    assert wb["segments_review"]["A3"].value == "review_0001"
