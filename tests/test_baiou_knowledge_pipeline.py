import json
from pathlib import Path

from openpyxl import load_workbook

from baiou.case_pipeline.knowledge import build_segment_assets
from baiou.case_pipeline.production.disabled_summary import write_disabled_summary


def write_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def make_case(batch_root: Path, case_id: str, segments: list[dict]) -> None:
    case_root = batch_root / "cases" / case_id
    write_json(case_root / "case_outline.json", {"case_summary": f"{case_id} summary", "stage_path": []})
    write_json(case_root / "segments.json", {"case_id": case_id, "schema_version": "segments_v01", "segments": segments})


def make_segment(segment_id: str, status: str, reply: str = "回复") -> dict:
    return {
        "case_id": "case_001",
        "segment_id": segment_id,
        "source_turn_ids": ["turn_0001"],
        "聊天阶段": "暧昧升温期",
        "女生状态": "热情",
        "男生目标": "延续话题",
        "推荐策略": "轻微调侃",
        "风险类型": [],
        "回复强度": "调侃",
        "当前上下文": "女生问男生在干嘛",
        "女生最后一句": "你在干嘛",
        "男生原回复": "想你",
        "原回复评价": "自然接话",
        "更优回复": reply,
        "迁移学习价值": "保持轻松",
        "quality_status": status,
        "human_review_applied": [{"choice": "暂不启用"}] if status == "disabled" else [],
    }


def read_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def test_disabled_summary_collects_disabled_segments(tmp_path: Path) -> None:
    batch_root = tmp_path / "segments" / "batch_a"
    make_case(batch_root, "case_001", [make_segment("seg_001", "approved"), make_segment("seg_002", "disabled")])

    summary = write_disabled_summary(batch_root)

    rows = read_jsonl(batch_root / "disabled_segments.jsonl")
    assert summary["disabled_count"] == 1
    assert rows[0]["segment_id"] == "seg_002"
    assert rows[0]["人工结论"] == "暂不启用"

    workbook = load_workbook(batch_root / "disabled_segments.xlsx")
    assert workbook["disabled_segments"].max_row == 2


def test_build_assets_maintains_current_knowledge_store(monkeypatch, tmp_path: Path) -> None:
    output_root = tmp_path / "cases"
    monkeypatch.setattr(build_segment_assets, "OUTPUT_ROOT", output_root)

    batch_a = output_root / "segments" / "batch_a"
    make_case(batch_a, "case_001", [make_segment("seg_001", "approved", "第一版"), make_segment("seg_002", "disabled")])
    first = build_segment_assets.build_assets("batch_a")

    current_root = output_root / "knowledge" / "current"
    assert first["segment_count"] == 1
    assert len(read_jsonl(current_root / "segments.jsonl")) == 1
    first_folders = [path for path in (current_root / "rag_knowledge_base" / "segments").iterdir() if path.is_dir()]
    assert len(first_folders) == 1
    assert len(list(first_folders[0].glob("*.md"))) == 1
    assert read_jsonl(output_root / "knowledge" / "imports" / "batch_a" / "skipped_segments.jsonl")[0]["segment_id"] == "seg_002"

    batch_b = output_root / "segments" / "batch_b"
    make_case(batch_b, "case_001", [make_segment("seg_001", "approved", "第二版"), make_segment("seg_003", "approved")])
    second = build_segment_assets.build_assets("batch_b")

    rows = read_jsonl(current_root / "segments.jsonl")
    by_id = {row["segment_id"]: row for row in rows}
    assert second["last_import"]["created_count"] == 1
    assert second["last_import"]["updated_count"] == 1
    assert len(rows) == 2
    assert by_id["seg_001"]["更优回复"] == "第二版"
    assert by_id["seg_001"]["rag_file_path"].startswith("segments/batch_b_")
    assert len(list((current_root / "rag_knowledge_base" / "segments").glob("*/*.md"))) == 3
    assert second["rag_knowledge_base"]["document_count"] == 2
    assert second["rag_knowledge_base"]["latest_import_document_count"] == 2
