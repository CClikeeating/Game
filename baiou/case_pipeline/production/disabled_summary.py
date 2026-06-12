from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from baiou.common.io import read_json, write_jsonl
from baiou.case_pipeline.knowledge.build_segment_index import case_dirs_for_batch


DISABLED_FIELDS = [
    "case_id",
    "segment_id",
    "source_turn_ids",
    "当前上下文",
    "女生最后一句",
    "男生原回复",
    "原回复评价",
    "更优回复",
    "迁移学习价值",
    "主标签",
    "次要标签",
    "quality_status",
    "禁用原因",
    "人工结论",
    "回复修正",
    "标签修正",
    "人工原则备注",
]


def write_disabled_summary(batch_root: Path) -> dict[str, Any]:
    rows = disabled_rows(batch_root)
    jsonl_path = batch_root / "disabled_segments.jsonl"
    xlsx_path = batch_root / "disabled_segments.xlsx"
    write_jsonl(jsonl_path, rows)
    write_disabled_workbook(xlsx_path, rows)
    return {"disabled_count": len(rows), "jsonl": str(jsonl_path), "xlsx": str(xlsx_path)}


def disabled_rows(batch_root: Path) -> list[dict[str, Any]]:
    cases_root = batch_root / "cases"
    if not cases_root.exists():
        return []
    rows: list[dict[str, Any]] = []
    for case_dir in case_dirs_for_batch(batch_root, cases_root):
        segments_path = case_dir / "segments.json"
        if not segments_path.exists():
            continue
        payload = read_json(segments_path)
        for segment in payload.get("segments", []) if isinstance(payload.get("segments", []), list) else []:
            if segment.get("quality_status") == "disabled":
                rows.append(disabled_row(segment))
    return rows


def disabled_row(segment: dict[str, Any]) -> dict[str, Any]:
    review = latest_human_review(segment)
    return {
        "case_id": segment.get("case_id", ""),
        "segment_id": segment.get("segment_id", ""),
        "source_turn_ids": ", ".join(str(item) for item in segment.get("source_turn_ids", []) if str(item)),
        "当前上下文": segment.get("当前上下文", ""),
        "女生最后一句": segment.get("女生最后一句", ""),
        "男生原回复": segment.get("男生原回复", ""),
        "原回复评价": segment.get("原回复评价", ""),
        "更优回复": segment.get("更优回复", ""),
        "迁移学习价值": segment.get("迁移学习价值", ""),
        "主标签": format_labels(segment),
        "次要标签": json.dumps(segment.get("次要标签", {}), ensure_ascii=False),
        "quality_status": segment.get("quality_status", ""),
        "禁用原因": disabled_reason(segment, review),
        "人工结论": review.get("choice", ""),
        "回复修正": review.get("reply_correction", ""),
        "标签修正": review.get("label_correction", ""),
        "人工原则备注": review.get("principle_note", ""),
    }


def latest_human_review(segment: dict[str, Any]) -> dict[str, Any]:
    reviews = segment.get("human_review_applied", [])
    if isinstance(reviews, list) and reviews and isinstance(reviews[-1], dict):
        return reviews[-1]
    return {}


def disabled_reason(segment: dict[str, Any], review: dict[str, Any]) -> str:
    notes = [str(review.get(key, "")).strip() for key in ["notes", "principle_note", "corrected_value"]]
    notes = [item for item in notes if item]
    if notes:
        return "\n".join(notes)
    reason = str(segment.get("quality_reason", "")).strip()
    if reason:
        return reason
    model_review = segment.get("model_review", {}) if isinstance(segment.get("model_review", {}), dict) else {}
    return str(model_review.get("reason", "")).strip()


def format_labels(segment: dict[str, Any]) -> str:
    risks = segment.get("风险类型", [])
    risk_text = "、".join(str(item) for item in risks if str(item)) if isinstance(risks, list) else str(risks or "")
    return "\n".join(
        [
            f"聊天阶段：{segment.get('聊天阶段', '')}",
            f"女生状态：{segment.get('女生状态', '')}",
            f"男生目标：{segment.get('男生目标', '')}",
            f"推荐策略：{segment.get('推荐策略', '')}",
            f"风险类型：{risk_text}",
            f"回复强度：{segment.get('回复强度', '')}",
        ]
    )


def write_disabled_workbook(path: Path, rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "disabled_segments"
    ws.append(DISABLED_FIELDS)
    for row in rows:
        ws.append([row.get(field, "") for field in DISABLED_FIELDS])
    header_fill = PatternFill("solid", fgColor="7F7F7F")
    header_font = Font(color="FFFFFF", bold=True)
    widths = {
        "case_id": 36,
        "segment_id": 42,
        "source_turn_ids": 36,
        "当前上下文": 90,
        "女生最后一句": 42,
        "男生原回复": 42,
        "原回复评价": 52,
        "更优回复": 42,
        "迁移学习价值": 54,
        "主标签": 30,
        "次要标签": 36,
        "禁用原因": 60,
        "人工结论": 18,
        "回复修正": 42,
        "标签修正": 42,
        "人工原则备注": 54,
    }
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for index, field in enumerate(DISABLED_FIELDS, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = widths.get(field, 22)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
