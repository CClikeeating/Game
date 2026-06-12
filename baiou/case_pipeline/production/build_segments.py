from __future__ import annotations

import argparse
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from baiou.common.io import PROJECT_ROOT, read_json, write_json
from baiou.case_pipeline.common import OUTPUT_ROOT, load_config, load_prompt
from baiou.common.chat_json_client import ChatJsonClient
from baiou.case_pipeline.schema import extract_transfer_value, normalize_segment, validate_segments
from baiou.case_pipeline.production.disabled_summary import write_disabled_summary

MAX_WORKERS_CAP = 50

REVIEW_FIELDS = [
    "review_id",
    "case_id",
    "原文连接/定位",
    "背景介绍",
    "当前上下文",
    "女生最后一句",
    "男生原回复",
    "主模型原回复评价",
    "主模型标签",
    "主模型建议回复",
    "主模型迁移学习价值",
    "主模型判断理由",
    "复核模型结论",
    "复核模型修改建议",
    "需要你复核的问题",
    "人工结论",
    "回复修正",
    "标签修正",
    "人工原则备注",
]

REVIEW_CHOICES = ["通过", "按复核模型修改", "手工修正", "拒绝", "暂不启用", "跳过"]
MISSING_NODE_FIELDS = [
    "missing_id",
    "case_id",
    "source_turn_ids",
    "原文连接/定位",
    "当前上下文",
    "复核模型漏拆理由",
    "优先级",
    "建议补拆重点",
    "人工结论",
    "人工备注",
]

MISSING_NODE_CHOICES = ["需要补拆", "不需要", "暂缓"]

DEFAULT_REVIEW_RULES = {
    "auto_approve_keep_original": {
        "enabled": True,
        "reply_prefixes": ["保留原回复"],
        "reason": "保留原回复且复核模型无修改意见",
    },
    "review_workbook": {
        "include_only_need_human_review": True,
        "review_choices": REVIEW_CHOICES,
        "missing_node_choices": MISSING_NODE_CHOICES,
    },
}


@dataclass(frozen=True)
class CaseJob:
    index: int
    case: dict[str, Any]
    user_id: str


def load_batch(input_bundle: str | Path) -> dict[str, Any]:
    bundle = Path(input_bundle)
    if not bundle.is_absolute():
        bundle = PROJECT_ROOT / bundle
    return read_json(bundle / "batch_chat_turns.json")


def find_cases(batch: dict[str, Any], case_ids: set[str] | None = None, limit: int | None = None) -> list[dict[str, Any]]:
    cases = [case for case in batch.get("cases", []) if not case_ids or case.get("case_id") in case_ids]
    return cases[:limit] if limit else cases


def load_run_options(max_workers: int | None = None, overwrite: bool = False) -> dict[str, Any]:
    options = load_config("run_options.json").get("case_plan", {})
    if max_workers is not None:
        options["max_workers"] = max_workers
    options["max_workers"] = max(1, min(MAX_WORKERS_CAP, int(options.get("max_workers", 1))))
    if overwrite:
        options["overwrite"] = True
    return options


def load_review_rules() -> dict[str, Any]:
    rules = json.loads(json.dumps(DEFAULT_REVIEW_RULES))
    loaded = load_config("review_rules.json")
    deep_update(rules, loaded)
    return rules


def deep_update(base: dict[str, Any], updates: dict[str, Any]) -> None:
    for key, value in updates.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            deep_update(base[key], value)
        else:
            base[key] = value


def build_jobs(cases: list[dict[str, Any]], models: dict[str, Any], options: dict[str, Any]) -> list[CaseJob]:
    max_workers = max(1, min(MAX_WORKERS_CAP, int(options.get("max_workers", 1))))
    strategy = str(options.get("user_id_strategy", "worker_index"))
    user_id_start = int(options.get("user_id_start", 1))
    prefix = str(options.get("user_id_prefix", ""))
    fixed_user_id = str(models.get("user_id", "71"))
    jobs: list[CaseJob] = []
    for index, case in enumerate(cases, start=1):
        if strategy == "fixed":
            user_id = fixed_user_id
        else:
            worker_slot = (index - 1) % max_workers
            user_id = f"{prefix}{user_id_start + worker_slot}"
        jobs.append(CaseJob(index=index, case=case, user_id=str(user_id)))
    return jobs


def compact_case(case: dict[str, Any], max_turns: int | None = None) -> dict[str, Any]:
    turns = []
    for block in case.get("blocks", []):
        for turn in block.get("turns", []):
            turns.append(
                {
                    "turn_id": turn.get("turn_id", ""),
                    "block_id": block.get("block_id", turn.get("source_block_id", "")),
                    "speaker": turn.get("speaker", ""),
                    "text": turn.get("text", ""),
                    "content_type": turn.get("content_type", "text"),
                    "time": turn.get("time", ""),
                    "source_image": turn.get("source_image", block.get("source_image", "")),
                }
            )
    if max_turns:
        turns = turns[:max_turns]
    return {"case_id": case.get("case_id", ""), "summary": case.get("summary", {}), "turns": turns}


def build_case_prompt(case: dict[str, Any]) -> str:
    taxonomy = load_config("taxonomy_v01.json")
    principles = load_config("prompt_principles.json")
    return "\n\n".join(
        [
            load_prompt("case_segment_v01.md"),
            "新版标签枚举：",
            json.dumps(taxonomy.get("labels", {}), ensure_ascii=False, indent=2),
            "Prompt 原则：",
            json.dumps(principles, ensure_ascii=False, indent=2),
            "完整聊天整案：",
            json.dumps(compact_case(case), ensure_ascii=False, indent=2),
        ]
    )


def build_review_prompt(case: dict[str, Any], primary: dict[str, Any]) -> str:
    taxonomy = load_config("taxonomy_v01.json")
    principles = load_config("prompt_principles.json")
    return "\n\n".join(
        [
            load_prompt("segment_review_v01.md"),
            "新版标签枚举：",
            json.dumps(taxonomy.get("labels", {}), ensure_ascii=False, indent=2),
            "Prompt 原则：",
            json.dumps(principles, ensure_ascii=False, indent=2),
            "完整聊天整案：",
            json.dumps(compact_case(case), ensure_ascii=False, indent=2),
            "主模型输出：",
            json.dumps(primary, ensure_ascii=False, indent=2),
        ]
    )


def run_segments(
    input_bundle: str,
    output_batch_id: str,
    case_ids: set[str] | None = None,
    case_limit: int | None = None,
    dry_run: bool = False,
    skip_review: bool = False,
    max_workers: int | None = None,
    overwrite: bool = False,
) -> dict[str, Any]:
    batch = load_batch(input_bundle)
    cases = find_cases(batch, case_ids, case_limit)
    output_root = OUTPUT_ROOT / "segments" / output_batch_id
    models = load_config("models.json")
    options = load_run_options(max_workers, overwrite)
    jobs = build_jobs(cases, models, options)
    prepare_output_dir(output_root, bool(options.get("overwrite", False)))

    if dry_run:
        previews = [
            {
                "case_index": job.index,
                "case_id": job.case.get("case_id", ""),
                "worker_user_id": job.user_id,
                "prompt": build_case_prompt(job.case)[:12000],
            }
            for job in jobs
        ]
        write_json(
            output_root / "dry_run_prompt_preview.json",
            {"cases": previews, "max_workers": options["max_workers"], "user_id_strategy": options.get("user_id_strategy", "")},
        )
        return {
            "status": "dry_run",
            "output_dir": str(output_root),
            "case_count": len(cases),
            "max_workers": options["max_workers"],
        }

    results: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=options["max_workers"]) as executor:
        futures = {executor.submit(process_job, job, output_root, models, skip_review): job for job in jobs}
        for future in as_completed(futures):
            job = futures[future]
            case_id = str(job.case.get("case_id", ""))
            try:
                result = future.result()
            except Exception as exc:  # noqa: BLE001 - keep batch processing and report failed cases.
                errors.append(
                    {
                        "case_index": job.index,
                        "case_id": case_id,
                        "worker_user_id": job.user_id,
                        "status": "failed",
                        "error": f"{exc.__class__.__name__}: {exc}",
                    }
                )
                if options.get("fail_fast", False):
                    raise
                continue
            results.append(result)
            print(
                json.dumps(
                    {
                        "case_index": result["case_index"],
                        "case_id": result["manifest_row"]["case_id"],
                        "user_id": result["manifest_row"]["worker_user_id"],
                        "segment_count": result["manifest_row"]["segment_count"],
                        "elapsed_seconds": result["manifest_row"]["elapsed_seconds"],
                    },
                    ensure_ascii=False,
                )
            )

    results.sort(key=lambda item: item.get("case_index", 0))
    manifest = [item["manifest_row"] for item in results]
    logs = [log for item in results for log in item["model_log"]]
    review_rows: list[dict[str, Any]] = []
    missing_node_rows: list[dict[str, Any]] = []
    for item in results:
        case_id = str(item["manifest_row"].get("case_id", ""))
        review = item.get("review", {})
        case = item.get("case", {})
        review_rows.extend(
            rows_for_review(
                case_id,
                item.get("segments", []),
                review,
                len(review_rows) + 1,
                case,
            )
        )
        missing_node_rows.extend(rows_for_missing_nodes(case_id, review, len(missing_node_rows) + 1, case))

    write_json(
        output_root / "segments_manifest.json",
        {"batch_id": output_batch_id, "source_bundle": input_bundle, "case_count": len(manifest), "cases": manifest},
    )
    write_json(
        output_root / "model_call_log.json",
        {
            "batch_id": output_batch_id,
            "max_workers": options["max_workers"],
            "user_id_strategy": options.get("user_id_strategy", ""),
            "calls": logs,
            "errors": errors,
        },
    )
    write_json(
        output_root / "case_plan_run_log.json",
        {
            "batch_id": output_batch_id,
            "input_bundle": input_bundle,
            "max_workers": options["max_workers"],
            "selected_case_count": len(cases),
            "completed_case_count": len(manifest),
            "error_count": len(errors),
            "cases": [
                {"case_index": job.index, "case_id": job.case.get("case_id", ""), "worker_user_id": job.user_id}
                for job in jobs
            ],
            "errors": errors,
        },
    )
    write_review_workbook(output_root / "human_review_segments.xlsx", review_rows, missing_node_rows)
    disabled_summary = write_disabled_summary(output_root)
    write_json(
        output_root / "handoff_segments_v01.json",
        {"schema_version": "handoff_segments_v01", "batch_id": output_batch_id, "main_entry": "segments_manifest.json"},
    )
    return {
        "status": "ok",
        "output_dir": str(output_root),
        "case_count": len(manifest),
        "review_rows": len(review_rows),
        "disabled_count": disabled_summary["disabled_count"],
        "error_count": len(errors),
        "max_workers": options["max_workers"],
    }


def rebuild_review_workbook(batch_id: str, review_filename: str = "human_review_segments.xlsx") -> dict[str, Any]:
    batch_root = OUTPUT_ROOT / "segments" / batch_id
    manifest_path = batch_root / "segments_manifest.json"
    if not manifest_path.exists():
        raise FileNotFoundError(manifest_path)
    manifest = read_json(manifest_path)
    source_cases = source_cases_for_manifest(manifest)
    review_rows: list[dict[str, Any]] = []
    missing_node_rows: list[dict[str, Any]] = []
    for row in manifest.get("cases", []):
        case_id = str(row.get("case_id", ""))
        case_dir = Path(str(row.get("case_dir", ""))) if str(row.get("case_dir", "")).strip() else batch_root / "cases" / case_id
        if not case_dir.is_absolute():
            case_dir = batch_root / case_dir
        segments_path = case_dir / "segments.json"
        review_path = case_dir / "model_review.json"
        if not segments_path.exists():
            continue
        payload = read_json(segments_path)
        review = read_json(review_path) if review_path.exists() else {}
        case = source_cases.get(case_id, {})
        segments = payload.get("segments", []) if isinstance(payload.get("segments", []), list) else []
        changed = apply_review_policy(segments, bool(payload.get("validation_issues")))
        if changed:
            payload["segments"] = segments
            write_json(segments_path, payload)
        update_manifest_row_status(row, segments)
        review_rows.extend(rows_for_review(case_id, segments, review, len(review_rows) + 1, case))
        missing_node_rows.extend(rows_for_missing_nodes(case_id, review, len(missing_node_rows) + 1, case))
    write_json(manifest_path, manifest)
    workbook_path = batch_root / review_filename
    write_review_workbook(workbook_path, review_rows, missing_node_rows)
    disabled_summary = write_disabled_summary(batch_root)
    return {
        "batch_id": batch_id,
        "workbook": str(workbook_path),
        "review_rows": len(review_rows),
        "missing_node_rows": len(missing_node_rows),
        "disabled_count": disabled_summary["disabled_count"],
    }


def source_cases_for_manifest(manifest: dict[str, Any]) -> dict[str, dict[str, Any]]:
    source_bundle = str(manifest.get("source_bundle", "")).strip()
    if not source_bundle:
        return {}
    bundle = Path(source_bundle)
    if not bundle.is_absolute():
        bundle = PROJECT_ROOT / bundle
    batch_path = bundle / "batch_chat_turns.json"
    if not batch_path.exists():
        return {}
    batch = read_json(batch_path)
    return {str(case.get("case_id", "")): case for case in batch.get("cases", []) if isinstance(case, dict)}

def prepare_output_dir(output_root: Path, overwrite: bool) -> None:
    output_root.mkdir(parents=True, exist_ok=True)
    existing_markers = [
        output_root / "dry_run_prompt_preview.json",
        output_root / "segments_manifest.json",
        output_root / "model_call_log.json",
        output_root / "human_review_segments.xlsx",
    ]
    if not overwrite and any(path.exists() for path in existing_markers):
        raise FileExistsError(f"Output batch already exists: {output_root}. Use --overwrite or a new --output-batch-id.")


def process_job(job: CaseJob, output_root: Path, models: dict[str, Any], skip_review: bool) -> dict[str, Any]:
    started = time.time()
    case = job.case
    case_id = str(case.get("case_id", ""))
    case_dir = output_root / "cases" / case_id
    case_dir.mkdir(parents=True, exist_ok=True)

    primary_client = ChatJsonClient("case_primary", models["case_primary"], job.user_id)
    review_client = ChatJsonClient("case_review", models["case_review"], job.user_id)
    primary_result = primary_client.chat_json("只输出合法 JSON。", build_case_prompt(case))
    primary = primary_result.get("parsed", {}) if primary_result.get("status") == "model_success" else fallback_segments(case)
    review_result = {"status": "skipped", "parsed": {}, "user_id": job.user_id, "model": "", "provider": "", "elapsed_seconds": 0, "usage": {}}
    if not skip_review:
        review_result = review_client.chat_json("只输出合法 JSON。", build_review_prompt(case, primary))
    review = review_result.get("parsed", {}) if isinstance(review_result.get("parsed"), dict) else {}
    case_outline, segments = normalize_output(case_id, primary, review)
    issues = validate_segments(segments)
    apply_review_policy(segments, bool(issues))
    write_json(case_dir / "case_outline.json", case_outline)
    write_json(
        case_dir / "segments.json",
        {"case_id": case_id, "schema_version": "segments_v01", "segments": segments, "validation_issues": issues},
    )
    write_json(case_dir / "model_review.json", review)
    primary_ok = primary_result.get("status") == "model_success"
    review_ok = skip_review or review_result.get("status") in {"model_success", "skipped"}
    model_failed = not primary_ok or not review_ok
    need_review = any(item.get("need_human_review") for item in segments)
    status = "model_failed" if model_failed else "needs_review" if need_review or not segments else "ready"
    manifest_row = {
        "case_index": job.index,
        "case_id": case_id,
        "worker_user_id": job.user_id,
        "segment_count": len(segments),
        "need_review_count": sum(1 for item in segments if item.get("need_human_review")),
        "primary_model": primary_result.get("model", ""),
        "primary_status": primary_result.get("status", ""),
        "review_model": review_result.get("model", ""),
        "review_status": review_result.get("status", ""),
        "elapsed_seconds": round(time.time() - started, 2),
        "case_dir": str(case_dir),
        "status": status,
    }
    return {
        "case_index": job.index,
        "manifest_row": manifest_row,
        "segments": segments,
        "review": review,
        "case": case,
        "model_log": [
            compact_log(case_id, "primary", primary_result),
            compact_log(case_id, "review", review_result),
        ],
    }


def apply_review_policy(segments: list[dict[str, Any]], has_validation_issues: bool = False) -> bool:
    rules = load_review_rules()
    changed = False
    for segment in segments:
        before = json.dumps(segment, ensure_ascii=False, sort_keys=True)
        segment["迁移学习价值"] = str(segment.get("迁移学习价值") or extract_transfer_value(str(segment.get("quality_reason", ""))))
        if has_validation_issues:
            segment["need_human_review"] = True
            if segment.get("quality_status") in {"draft", "approved"}:
                segment["quality_status"] = "needs_review"
        elif is_auto_approved_keep_original(segment, rules):
            segment["quality_status"] = "approved"
            segment["need_human_review"] = False
            segment["auto_review_reason"] = str(rules["auto_approve_keep_original"].get("reason", "保留原回复且复核模型无修改意见"))
        elif segment.get("quality_status") in {"approved", "disabled", "source_error"}:
            segment["need_human_review"] = False
        elif segment.get("quality_status") == "rejected":
            segment["need_human_review"] = True
        else:
            segment["quality_status"] = "needs_review"
            segment["need_human_review"] = True
        after = json.dumps(segment, ensure_ascii=False, sort_keys=True)
        changed = changed or before != after
    return changed


def is_auto_approved_keep_original(segment: dict[str, Any], rules: dict[str, Any]) -> bool:
    config = rules.get("auto_approve_keep_original", {}) if isinstance(rules.get("auto_approve_keep_original", {}), dict) else {}
    if not config.get("enabled", True):
        return False
    reply = str(segment.get("更优回复", "")).strip()
    prefixes = config.get("reply_prefixes", ["保留原回复"])
    if not isinstance(prefixes, list):
        prefixes = ["保留原回复"]
    if not any(reply.startswith(str(prefix)) for prefix in prefixes if str(prefix)):
        return False
    return model_review_has_no_modifications(segment.get("model_review", {}))


def model_review_has_no_modifications(model_review: Any) -> bool:
    if not isinstance(model_review, dict) or not model_review:
        return True
    verdict = str(model_review.get("verdict", "")).strip().lower()
    if verdict in {"revise", "reject"}:
        return False
    issues = model_review.get("issues", [])
    if isinstance(issues, list) and issues:
        return False
    return verdict in {"", "pass", "agree"}


def update_manifest_row_status(row: dict[str, Any], segments: list[dict[str, Any]]) -> None:
    row["segment_count"] = len(segments)
    row["need_review_count"] = sum(1 for item in segments if item.get("need_human_review"))
    model_failed = row.get("primary_status") != "model_success" or row.get("review_status") not in {"model_success", "skipped"}
    row["status"] = "model_failed" if model_failed else "needs_review" if row["need_review_count"] or not segments else "ready"


def normalize_output(case_id: str, primary: dict[str, Any], review: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    outline = primary.get("case_outline", {}) if isinstance(primary.get("case_outline"), dict) else {}
    outline = {
        "case_id": str(outline.get("case_id") or case_id),
        "schema_version": "segments_v01",
        "case_summary": str(outline.get("case_summary", "")),
        "stage_path": outline.get("stage_path", []) if isinstance(outline.get("stage_path"), list) else [],
        "key_nodes": outline.get("key_nodes", []) if isinstance(outline.get("key_nodes"), list) else [],
        "final_outcome": str(outline.get("final_outcome", "")),
        "usable_for_mvp": bool(outline.get("usable_for_mvp", True)),
        "notes": str(outline.get("notes", "")),
    }
    review_by_id = {item.get("segment_id", ""): item for item in review.get("segment_reviews", []) if isinstance(item, dict)}
    segments = []
    for index, raw_segment in enumerate(primary.get("segments", []) if isinstance(primary.get("segments"), list) else [], start=1):
        if not isinstance(raw_segment, dict):
            continue
        raw_segment_id = str(raw_segment.get("segment_id", ""))
        segment = normalize_segment(raw_segment, case_id, index)
        item_review = review_by_id.get(segment["segment_id"], {}) or review_by_id.get(raw_segment_id, {})
        if item_review:
            segment["model_review"] = item_review
            if item_review.get("verdict") in {"revise", "reject"}:
                segment["need_human_review"] = True
                segment["quality_status"] = "needs_review" if item_review.get("verdict") == "revise" else "rejected"
        segments.append(segment)
    return outline, segments


def fallback_segments(case: dict[str, Any]) -> dict[str, Any]:
    return {
        "case_outline": {
            "case_id": case.get("case_id", ""),
            "schema_version": "segments_v01",
            "case_summary": "模型未生成，需人工处理。",
            "stage_path": [],
            "key_nodes": [],
            "final_outcome": "",
            "usable_for_mvp": False,
            "notes": "primary model failed",
        },
        "segments": [],
    }


def rows_for_review(
    case_id: str,
    segments: list[dict[str, Any]],
    review: dict[str, Any],
    start: int,
    case: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    case = case or {}
    rules = load_review_rules()
    workbook_config = rules.get("review_workbook", {}) if isinstance(rules.get("review_workbook", {}), dict) else {}
    include_only_need_review = bool(workbook_config.get("include_only_need_human_review", True))
    rows = []
    for segment in segments:
        if include_only_need_review and not segment.get("need_human_review"):
            continue
        rows.append(
            {
                "review_id": f"review_{start + len(rows):04d}",
                "case_id": case_id,
                "segment_id": segment.get("segment_id", ""),
                "原文连接/定位": source_location(case, segment),
                "背景介绍": segment.get("当前上下文", ""),
                "当前上下文": surrounding_context(case, segment.get("source_turn_ids", []), window=10),
                "女生最后一句": segment.get("女生最后一句", ""),
                "男生原回复": segment.get("男生原回复", ""),
                "主模型原回复评价": segment.get("原回复评价", ""),
                "主模型标签": format_labels(segment),
                "主模型建议回复": segment.get("更优回复", ""),
                "主模型迁移学习价值": segment.get("迁移学习价值", ""),
                "主模型判断理由": segment.get("quality_reason", ""),
                "复核模型结论": review_verdict_text(segment.get("model_review", {})),
                "复核模型修改建议": review_issues_text(segment.get("model_review", {})),
                "需要你复核的问题": review_task_text(segment),
                "人工结论": "",
                "回复修正": "",
                "标签修正": "",
                "人工原则备注": "",
            }
        )
    return rows

def rows_for_missing_nodes(
    case_id: str,
    review: dict[str, Any],
    start: int,
    case: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    case = case or {}
    raw_nodes = review.get("missing_nodes", []) if isinstance(review, dict) else []
    if not isinstance(raw_nodes, list):
        return []
    rows: list[dict[str, Any]] = []
    for offset, node in enumerate(raw_nodes):
        if not isinstance(node, dict):
            continue
        source_turn_ids = node.get("source_turn_ids", [])
        if isinstance(source_turn_ids, str):
            source_turn_ids = [source_turn_ids] if source_turn_ids else []
        elif not isinstance(source_turn_ids, list):
            source_turn_ids = []
        locator_segment = {"source_turn_ids": source_turn_ids}
        rows.append(
            {
                "missing_id": f"missing_{start + offset:04d}",
                "case_id": case_id,
                "source_turn_ids": ", ".join(str(item) for item in source_turn_ids if str(item)),
                "原文连接/定位": source_location(case, locator_segment),
                "当前上下文": surrounding_context(case, source_turn_ids, window=10),
                "复核模型漏拆理由": node.get("reason", ""),
                "优先级": node.get("priority", ""),
                "建议补拆重点": node.get("suggested_segment_focus", ""),
                "人工结论": "",
                "人工备注": "",
            }
        )
    return rows

def format_labels(segment: dict[str, Any]) -> str:
    risks = segment.get("风险类型", [])
    risk_text = "、".join(risks) if isinstance(risks, list) else str(risks or "")
    parts = [
        f"聊天阶段：{segment.get('聊天阶段', '')}",
        f"女生状态：{segment.get('女生状态', '')}",
        f"男生目标：{segment.get('男生目标', '')}",
        f"推荐策略：{segment.get('推荐策略', '')}",
        f"风险类型：{risk_text}",
        f"回复强度：{segment.get('回复强度', '')}",
    ]
    secondary = format_secondary_labels(segment.get("次要标签", {}))
    if secondary:
        parts.extend(["", "次要标签：", secondary])
    return "\n".join(parts)


def format_secondary_labels(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    lines = []
    for field in ["聊天阶段", "女生状态", "男生目标", "推荐策略", "回复强度"]:
        item = str(value.get(field, "")).strip()
        if item:
            lines.append(f"{field}：{item}")
    risks = value.get("风险类型", [])
    if isinstance(risks, list) and risks:
        lines.append("风险类型：" + "、".join(str(item) for item in risks if str(item)))
    note = str(value.get("说明", "")).strip()
    if note:
        lines.append(f"说明：{note}")
    return "\n".join(lines)


def review_verdict_text(model_review: Any) -> str:
    if not isinstance(model_review, dict) or not model_review:
        return "复核模型未给出单独意见。"
    verdict = model_review.get("verdict", "")
    need = model_review.get("need_human_review", "")
    verdict_cn = {"pass": "通过", "revise": "建议修改", "reject": "建议拒绝"}.get(str(verdict), str(verdict))
    return f"复核结论：{verdict_cn}\n复核模型认为是否需要人工复核：{need}"


def review_issues_text(model_review: Any) -> str:
    if not isinstance(model_review, dict):
        return ""
    issues = model_review.get("issues", []) if isinstance(model_review.get("issues", []), list) else []
    if not issues:
        return "无修改建议。"
    lines = []
    for index, issue in enumerate(issues, start=1):
        if not isinstance(issue, dict):
            continue
        lines.append(
            f"{index}. 字段：{issue.get('field', '')}\n"
            f"   主模型当前值：{stringify_review_value(issue.get('current_value', ''))}\n"
            f"   复核建议值：{stringify_review_value(issue.get('suggested_value', ''))}\n"
            f"   理由：{issue.get('reason', '')}"
        )
    return "\n".join(lines)


def stringify_review_value(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def review_task_text(segment: dict[str, Any]) -> str:
    return (
        "请对照原文定位和当前上下文，判断主模型拆出的片段是否可入库。\n"
        "通过=认可主模型当前字段；按复核模型修改=采用复核模型修改建议；手工修正=在人工修正列写 JSON 或直接写更合适回复；"
        "拒绝=该片段不用；暂不启用=保留但不入库；跳过=这次先不处理。"
    )


def flatten_case_turns(case: dict[str, Any]) -> list[dict[str, Any]]:
    turns = []
    for block in case.get("blocks", []):
        for turn in block.get("turns", []) if isinstance(block.get("turns", []), list) else []:
            item = dict(turn)
            item.setdefault("block_id", block.get("block_id", item.get("source_block_id", "")))
            item.setdefault("source_image", turn.get("source_image", block.get("source_image", "")))
            turns.append(item)
    return turns


def source_location(case: dict[str, Any], segment: dict[str, Any]) -> str:
    turn_ids = [str(item) for item in segment.get("source_turn_ids", []) if str(item)]
    turns = flatten_case_turns(case)
    by_id = {str(turn.get("turn_id", "")): turn for turn in turns}
    images = []
    for turn_id in turn_ids:
        image = str(by_id.get(turn_id, {}).get("source_image", "")).strip()
        if image and image not in images:
            images.append(image)
    lines = []
    if turn_ids:
        lines.append("turn_ids: " + ", ".join(turn_ids))
    lines.extend("image: " + image for image in images)
    return "\n".join(lines)


def surrounding_context(case: dict[str, Any], source_turn_ids: list[Any], window: int = 10) -> str:
    turns = flatten_case_turns(case)
    if not turns:
        return ""
    wanted = {str(item) for item in source_turn_ids if str(item)}
    positions = [index for index, turn in enumerate(turns) if str(turn.get("turn_id", "")) in wanted]
    if not positions:
        return ""
    start = max(0, min(positions) - window)
    end = min(len(turns), max(positions) + window + 1)
    lines = []
    for turn in turns[start:end]:
        marker = "*" if str(turn.get("turn_id", "")) in wanted else " "
        speaker = speaker_cn(str(turn.get("speaker", "")))
        text = str(turn.get("text", "")).strip()
        visual = str(turn.get("visual_note", "")).strip()
        suffix = f"（{visual}）" if visual else ""
        lines.append(f"{marker} {turn.get('turn_id', '')} {speaker}: {text}{suffix}")
    return "\n".join(lines)


def speaker_cn(value: str) -> str:
    return {"male": "男生", "female": "女生", "system": "系统", "narration": "旁白"}.get(value, value or "未知")

def write_review_workbook(path: Path, rows: list[dict[str, Any]], missing_rows: list[dict[str, Any]] | None = None) -> None:
    rules = load_review_rules()
    workbook_config = rules.get("review_workbook", {}) if isinstance(rules.get("review_workbook", {}), dict) else {}
    review_choices = workbook_config.get("review_choices", REVIEW_CHOICES)
    if not isinstance(review_choices, list):
        review_choices = REVIEW_CHOICES
    missing_node_choices = workbook_config.get("missing_node_choices", MISSING_NODE_CHOICES)
    if not isinstance(missing_node_choices, list):
        missing_node_choices = MISSING_NODE_CHOICES
    wb = Workbook()
    ws = wb.active
    ws.title = "segments_review"
    write_sheet(
        ws,
        REVIEW_FIELDS,
        rows,
        {
            "review_id": 14,
            "case_id": 38,
            "原文连接/定位": 70,
            "背景介绍": 48,
            "当前上下文": 90,
            "女生最后一句": 46,
            "男生原回复": 46,
            "主模型原回复评价": 52,
            "主模型标签": 30,
            "主模型建议回复": 46,
            "主模型迁移学习价值": 54,
            "主模型判断理由": 52,
            "复核模型结论": 28,
            "复核模型修改建议": 70,
            "需要你复核的问题": 70,
            "人工结论": 24,
            "回复修正": 54,
            "标签修正": 54,
            "人工原则备注": 60,
        },
        [str(item) for item in review_choices],
    )
    write_review_meta_sheet(wb, rows)
    missing_ws = wb.create_sheet("missing_nodes_review")
    write_sheet(
        missing_ws,
        MISSING_NODE_FIELDS,
        missing_rows or [],
        {
            "missing_id": 16,
            "case_id": 38,
            "source_turn_ids": 42,
            "原文连接/定位": 70,
            "当前上下文": 90,
            "复核模型漏拆理由": 60,
            "优先级": 16,
            "建议补拆重点": 60,
            "人工结论": 24,
            "人工备注": 50,
        },
        [str(item) for item in missing_node_choices],
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_review_meta_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("_review_meta")
    ws.sheet_state = "hidden"
    ws.append(["sheet", "review_id", "case_id", "segment_id"])
    for row in rows:
        ws.append(["segments_review", row.get("review_id", ""), row.get("case_id", ""), row.get("segment_id", "")])


def write_sheet(
    ws: Any,
    fields: list[str],
    rows: list[dict[str, Any]],
    field_widths: dict[str, int],
    choices: list[str] | None = None,
) -> None:
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for index, field in enumerate(fields, start=1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = field_widths.get(field, 24)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if rows and choices and "人工结论" in fields:
        validation = DataValidation(type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True)
        ws.add_data_validation(validation)
        choice_col = fields.index("人工结论") + 1
        validation.add(f"{ws.cell(2, choice_col).coordinate}:{ws.cell(len(rows) + 1, choice_col).coordinate}")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def compact_log(case_id: str, role: str, result: dict[str, Any]) -> dict[str, Any]:
    return {
        "case_id": case_id,
        "role": role,
        "provider": result.get("provider", ""),
        "model": result.get("model", ""),
        "user_id": result.get("user_id", ""),
        "status": result.get("status", ""),
        "error": result.get("error", ""),
        "elapsed_seconds": result.get("elapsed_seconds", 0),
        "usage": result.get("usage", {}),
        "raw_text_preview": str(result.get("raw_text", ""))[:4000],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Baiou segments_v01 from source pipeline batch output.")
    parser.add_argument("--input-bundle")
    parser.add_argument("--output-batch-id", required=True)
    parser.add_argument("--case-id", action="append", default=[])
    parser.add_argument("--case-limit", type=int)
    parser.add_argument("--max-workers", type=int)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-review", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--rebuild-review-workbook", action="store_true")
    args = parser.parse_args()
    if args.rebuild_review_workbook:
        result = rebuild_review_workbook(args.output_batch_id)
    else:
        if not args.input_bundle:
            parser.error("--input-bundle is required unless --rebuild-review-workbook is used")
        result = run_segments(
            input_bundle=args.input_bundle,
            output_batch_id=args.output_batch_id,
            case_ids=set(args.case_id) if args.case_id else None,
            case_limit=args.case_limit,
            dry_run=args.dry_run,
            skip_review=args.skip_review,
            max_workers=args.max_workers,
            overwrite=args.overwrite,
        )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()


