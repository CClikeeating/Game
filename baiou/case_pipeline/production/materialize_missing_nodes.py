from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from baiou.common.io import PROJECT_ROOT, read_json, write_json
from baiou.case_pipeline.common import OUTPUT_ROOT
from baiou.case_pipeline.production.apply_review import load_source_cases, source_turns
from baiou.case_pipeline.schema import normalize_segment


DEFAULT_CHOICE = "需要补拆"
DEFAULT_SEGMENT_PREFIX = "supp_missing"


def materialize_missing_nodes(
    batch_id: str,
    review_xlsx: str | Path | None = None,
    segments_root: str | Path | None = None,
    choice: str = DEFAULT_CHOICE,
    segment_prefix: str = DEFAULT_SEGMENT_PREFIX,
    dry_run: bool = False,
) -> dict[str, Any]:
    batch_root = resolve_batch_root(batch_id, segments_root)
    workbook_path = Path(review_xlsx) if review_xlsx else batch_root / "human_review_segments.xlsx"
    rows = [row for row in rows_from_missing_nodes(workbook_path) if str(row.get("人工结论") or "").strip() == choice]
    source_cases = load_source_cases(batch_root)

    created = 0
    updated = 0
    skipped: list[dict[str, Any]] = []
    touched_cases: set[str] = set()
    materialized: list[dict[str, Any]] = []

    for row in rows:
        case_id = str(row.get("case_id") or "").strip()
        missing_id = str(row.get("missing_id") or "").strip()
        if not case_id or not missing_id:
            skipped.append({"case_id": case_id, "missing_id": missing_id, "reason": "missing_identity"})
            continue
        case_dir = batch_root / "cases" / case_id
        segments_path = case_dir / "segments.json"
        if not segments_path.exists():
            skipped.append({"case_id": case_id, "missing_id": missing_id, "reason": "segments_not_found"})
            continue
        source_case = source_cases.get(case_id, {})
        turn_ids = parse_turn_ids(row.get("source_turn_ids"))
        turns = source_turns(source_case, turn_ids) if source_case else []
        if not turns:
            skipped.append({"case_id": case_id, "missing_id": missing_id, "reason": "source_turns_not_found"})
            continue

        payload = read_json(segments_path)
        segments = payload.get("segments", []) if isinstance(payload.get("segments"), list) else []
        segment_id = f"{segment_prefix}_{missing_id}"
        normalized_id = f"{case_id}_{segment_id}"
        existed = any(str(item.get("segment_id")) == normalized_id for item in segments if isinstance(item, dict))
        segment = build_supplemental_segment(case_id, segment_id, row, turn_ids, turns)
        next_segments = [item for item in segments if not (isinstance(item, dict) and str(item.get("segment_id")) == normalized_id)]
        next_segments.append(segment)
        payload["segments"] = next_segments

        if not dry_run:
            write_json(segments_path, payload)
        touched_cases.add(case_id)
        materialized.append({"case_id": case_id, "missing_id": missing_id, "segment_id": normalized_id, "updated": existed})
        if existed:
            updated += 1
        else:
            created += 1

    summary = {
        "schema_version": "materialized_missing_nodes_v01",
        "batch_id": batch_id,
        "review_xlsx": str(workbook_path),
        "choice": choice,
        "dry_run": dry_run,
        "created_count": created,
        "updated_count": updated,
        "skipped_count": len(skipped),
        "touched_cases": sorted(touched_cases),
        "materialized": materialized,
        "skipped": skipped,
    }
    if not dry_run:
        write_json(batch_root / "materialized_missing_nodes_summary.json", summary)
        refresh_manifest_counts(batch_root)
    return summary


def build_supplemental_segment(
    case_id: str,
    segment_id: str,
    row: dict[str, Any],
    turn_ids: list[str],
    turns: list[dict[str, Any]],
) -> dict[str, Any]:
    focus = str(row.get("建议补拆重点") or "").strip()
    reason = str(row.get("复核模型漏拆理由") or "").strip()
    note = str(row.get("人工备注") or "").strip()
    labels = infer_labels(focus, reason, note)
    female_last, male_reply = infer_reply_pair(turns)
    context = str(row.get("当前上下文") or "").strip() or render_turns(turns)
    transfer_value = build_transfer_value(focus, reason, note)
    segment = {
        "case_id": case_id,
        "schema_version": "segments_v01",
        "segment_id": segment_id,
        "source_turn_ids": turn_ids,
        "当前上下文": context,
        "女生最后一句": female_last,
        "男生原回复": male_reply,
        "原回复评价": "有效：人工确认需要补拆；作为边界/校准样本入库，重点是学习低压力承接、自然收住或框架处理。",
        "聊天阶段": labels["聊天阶段"],
        "接触状态": labels["接触状态"],
        "关系推进目标": labels["关系推进目标"],
        "女生状态": labels["女生状态"],
        "男生目标": labels["男生目标"],
        "推荐策略": labels["推荐策略"],
        "风险类型": [],
        "回复强度": labels["回复强度"],
        "高热度信号": labels["高热度信号"],
        "次要标签": {
            "说明": "由人工确认的 missing_nodes_review 补拆生成；标签按保守启发式初标，可在后续批量复核中再精修。"
        },
        "更优回复": male_reply,
        "迁移学习价值": transfer_value,
        "quality_status": "approved",
        "need_human_review": False,
        "quality_reason": f"人工漏拆补片：{transfer_value}",
        "created_from_missing_node": {
            "missing_id": row.get("missing_id", ""),
            "choice": row.get("人工结论", ""),
            "reason": reason,
            "focus": focus,
            "note": note,
        },
    }
    return normalize_segment(segment, case_id, 0)


def infer_labels(*texts: str) -> dict[str, str]:
    text = "\n".join(item for item in texts if item)
    labels = {
        "聊天阶段": "熟悉期",
        "接触状态": "未知",
        "关系推进目标": "无",
        "女生状态": "正常",
        "男生目标": "接话",
        "推荐策略": "主动降压",
        "回复强度": "安全",
        "高热度信号": "无",
    }
    if has_any(text, ["拒绝", "后撤", "软抗拒", "敏感", "测试", "抱怨", "不适合", "自然收尾", "弱承接", "低压力", "降压"]):
        labels.update({"关系推进目标": "降压修复", "女生状态": "防御", "男生目标": "降压", "推荐策略": "主动降压", "回复强度": "安全"})
    if has_any(text, ["好感", "推拉", "增加投入", "暧昧"]):
        labels.update({"聊天阶段": "暧昧升温期", "关系推进目标": "暧昧升温", "女生状态": "热情", "男生目标": "升温", "推荐策略": "轻微调侃", "回复强度": "调侃", "高热度信号": "暧昧试探"})
    if has_any(text, ["邀约", "见面"]):
        labels.update({"关系推进目标": "邀约见面", "男生目标": "邀约"})
    if has_any(text, ["自然收尾", "弱承接", "低压力", "降压", "关心", "晚安"]):
        labels.update({"高热度信号": "无", "回复强度": "安全", "推荐策略": "主动降压"})
    return labels


def infer_reply_pair(turns: list[dict[str, Any]]) -> tuple[str, str]:
    if not turns:
        return "", ""
    male_indexes = [index for index, turn in enumerate(turns) if is_male(turn)]
    female_indexes = [index for index, turn in enumerate(turns) if is_female(turn)]
    if not male_indexes:
        return last_text(turns, female_indexes), ""
    male_index = male_indexes[-1]
    female_before = [index for index in female_indexes if index < male_index]
    if female_before:
        female_index = female_before[-1]
        male_reply = collect_consecutive_text(turns, female_index + 1, "male")
        return str(turns[female_index].get("text") or ""), male_reply or str(turns[male_index].get("text") or "")
    female_after = [index for index in female_indexes if index > male_index]
    if female_after:
        female_index = female_after[-1]
        male_reply = collect_consecutive_text(turns, max(0, male_index - 1), "male") or str(turns[male_index].get("text") or "")
        return str(turns[female_index].get("text") or ""), male_reply
    return "", str(turns[male_index].get("text") or "")


def collect_consecutive_text(turns: list[dict[str, Any]], start: int, speaker: str) -> str:
    chunks: list[str] = []
    for turn in turns[start:]:
        if speaker == "male" and not is_male(turn):
            if chunks:
                break
            continue
        if speaker == "female" and not is_female(turn):
            if chunks:
                break
            continue
        text = str(turn.get("text") or "").strip()
        if text:
            chunks.append(text)
    return " / ".join(chunks)


def last_text(turns: list[dict[str, Any]], indexes: list[int]) -> str:
    return str(turns[indexes[-1]].get("text") or "") if indexes else ""


def is_male(turn: dict[str, Any]) -> bool:
    speaker = str(turn.get("speaker") or "").lower()
    return "male" in speaker or "男" in speaker


def is_female(turn: dict[str, Any]) -> bool:
    speaker = str(turn.get("speaker") or "").lower()
    return "female" in speaker or "女" in speaker


def build_transfer_value(focus: str, reason: str, note: str) -> str:
    chunks = [item for item in [focus, reason, note] if item]
    return "；".join(chunks) if chunks else "补充人工确认的低压力/弱承接/自然收尾校准节点。"


def render_turns(turns: list[dict[str, Any]]) -> str:
    lines = []
    for turn in turns:
        speaker = str(turn.get("speaker") or "").strip()
        text = str(turn.get("text") or "").strip()
        turn_id = str(turn.get("turn_id") or "").strip()
        lines.append(f"{turn_id} {speaker}: {text}".strip())
    return "\n".join(lines)


def rows_from_missing_nodes(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "missing_nodes_review" not in wb.sheetnames:
        return []
    ws = wb["missing_nodes_review"]
    headers = [str(cell.value or "") for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def parse_turn_ids(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in re.split(r"[,，\n]+", str(value or "")) if item.strip()]


def has_any(text: str, needles: list[str]) -> bool:
    return any(needle in text for needle in needles)


def resolve_batch_root(batch_id: str, segments_root: str | Path | None) -> Path:
    root = Path(segments_root) if segments_root else OUTPUT_ROOT / "segments"
    root = root if root.is_absolute() else PROJECT_ROOT / root
    return root / batch_id


def refresh_manifest_counts(batch_root: Path) -> None:
    manifest_path = batch_root / "segments_manifest.json"
    if not manifest_path.exists():
        return
    manifest = read_json(manifest_path)
    for row in manifest.get("cases", []) if isinstance(manifest.get("cases"), list) else []:
        case_id = str(row.get("case_id") or "")
        case_path = batch_root / "cases" / case_id / "segments.json"
        if not case_path.exists():
            continue
        payload = read_json(case_path)
        segments = payload.get("segments", []) if isinstance(payload.get("segments"), list) else []
        row["segment_count"] = len(segments)
        row["approved_count"] = sum(1 for item in segments if isinstance(item, dict) and item.get("quality_status") == "approved")
        row["need_review_count"] = sum(1 for item in segments if isinstance(item, dict) and item.get("need_human_review"))
        row["status"] = "ready" if row["approved_count"] else "needs_review"
    write_json(manifest_path, manifest)


def main() -> None:
    parser = argparse.ArgumentParser(description="Turn reviewed missing_nodes_review rows into approved supplemental segments.")
    parser.add_argument("--batch-id", required=True)
    parser.add_argument("--review-xlsx")
    parser.add_argument("--segments-root")
    parser.add_argument("--choice", default=DEFAULT_CHOICE)
    parser.add_argument("--segment-prefix", default=DEFAULT_SEGMENT_PREFIX)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    print(
        json.dumps(
            materialize_missing_nodes(
                batch_id=args.batch_id,
                review_xlsx=args.review_xlsx,
                segments_root=args.segments_root,
                choice=args.choice,
                segment_prefix=args.segment_prefix,
                dry_run=args.dry_run,
            ),
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
