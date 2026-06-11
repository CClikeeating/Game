from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from workflow.common.io import PROJECT_ROOT, read_json, write_json
from workV.common import OUTPUT_ROOT

CORRECTABLE_FIELDS = [
    "当前上下文",
    "女生最后一句",
    "男生原回复",
    "原回复评价",
    "聊天阶段",
    "女生状态",
    "男生目标",
    "推荐策略",
    "风险类型",
    "回复强度",
    "次要标签",
    "更优回复",
    "下一步建议",
]


def apply_review(batch_id: str, review_path: str | None = None) -> dict[str, Any]:
    batch_root = OUTPUT_ROOT / "segments" / batch_id
    path = Path(review_path) if review_path else batch_root / "human_review_segments.xlsx"
    rows = rows_from_workbook(path)
    touched: set[str] = set()
    applied = 0
    source_fix_items: list[dict[str, Any]] = []
    source_cases = load_source_cases(batch_root)
    by_case: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        if user_choice(row) or user_correction(row) or user_notes(row):
            by_case.setdefault(str(row.get("case_id", "")), []).append(row)
    for case_id, case_rows in by_case.items():
        case_path = batch_root / "cases" / case_id / "segments.json"
        if not case_path.exists():
            continue
        payload = read_json(case_path)
        segments = payload.get("segments", [])
        by_segment = {segment.get("segment_id", ""): segment for segment in segments}
        for row in case_rows:
            segment = by_segment.get(str(row.get("segment_id", "")))
            if not segment:
                continue
            choice = user_choice(row)
            corrected = user_correction(row)
            if choice == "通过":
                segment["quality_status"] = "approved"
                segment["need_human_review"] = False
            elif choice == "按复核模型修改":
                apply_model_review_suggestions(segment)
                segment["quality_status"] = "approved"
                segment["need_human_review"] = False
            elif choice == "拒绝":
                segment["quality_status"] = "rejected"
                segment["need_human_review"] = False
            elif choice == "暂不启用":
                segment["quality_status"] = "disabled"
                segment["need_human_review"] = False
            elif choice == "说话人错误，回源修正":
                segment["quality_status"] = "source_error"
                segment["need_human_review"] = False
                source_fix_items.append(source_fix_item(case_id, segment, row, source_cases.get(case_id, {})))
            elif choice == "手工修正" or corrected:
                applied_patch = apply_corrections(segment, row, corrected)
                if applied_patch:
                    segment["quality_status"] = "approved"
                    segment["need_human_review"] = False
                else:
                    segment["quality_status"] = "needs_review"
                    segment["need_human_review"] = True
            elif choice == "跳过":
                segment["quality_status"] = segment.get("quality_status", "draft")
            else:
                continue
            segment.setdefault("human_review_applied", []).append(
                {
                    "review_id": row.get("review_id", ""),
                    "choice": choice,
                    "corrected_value": corrected,
                    "notes": user_notes(row),
                }
            )
            applied += 1
        write_json(case_path, payload)
        touched.add(case_id)
    if source_fix_items:
        write_source_fix_queue(batch_root, source_fix_items)
    refresh_manifest(batch_root)
    return {
        "batch_id": batch_id,
        "rows_with_input": len(rows),
        "applied": applied,
        "source_fix_count": len(source_fix_items),
        "touched_cases": sorted(touched),
    }



def load_source_cases(batch_root: Path) -> dict[str, dict[str, Any]]:
    manifest_path = batch_root / "segments_manifest.json"
    if not manifest_path.exists():
        return {}
    manifest = read_json(manifest_path)
    source_bundle = str(manifest.get("source_bundle", "")).strip()
    if not source_bundle:
        return {}
    bundle = Path(source_bundle)
    if not bundle.is_absolute():
        bundle = PROJECT_ROOT / bundle
    batch_path = bundle / "batch_chat_turns.json"
    if not batch_path.exists():
        return {}
    batch = read_json(batch_path)
    return {str(case.get("case_id", "")): case for case in batch.get("cases", []) if isinstance(case, dict)}


def source_fix_item(case_id: str, segment: dict[str, Any], row: dict[str, Any], case: dict[str, Any]) -> dict[str, Any]:
    turn_ids = [str(item) for item in segment.get("source_turn_ids", []) if str(item)]
    turns = source_turns(case, turn_ids)
    return {
        "status": "pending",
        "case_id": case_id,
        "segment_id": segment.get("segment_id", ""),
        "source_turn_ids": turn_ids,
        "turns": turns,
        "source_location": row.get("原文连接/定位", ""),
        "human_notes": user_notes(row) or user_correction(row),
        "created_by": "apply_review",
    }


def source_turns(case: dict[str, Any], turn_ids: list[str]) -> list[dict[str, Any]]:
    wanted = set(turn_ids)
    rows: list[dict[str, Any]] = []
    for block in case.get("blocks", []) if isinstance(case.get("blocks", []), list) else []:
        for turn in block.get("turns", []) if isinstance(block.get("turns", []), list) else []:
            turn_id = str(turn.get("turn_id", ""))
            if turn_id not in wanted:
                continue
            rows.append(
                {
                    "turn_id": turn_id,
                    "speaker": turn.get("speaker", ""),
                    "text": turn.get("text", ""),
                    "content_type": turn.get("content_type", ""),
                    "time": turn.get("time", ""),
                    "source_image": turn.get("source_image", block.get("source_image", "")),
                    "block_id": block.get("block_id", turn.get("source_block_id", "")),
                }
            )
    return rows


def write_source_fix_queue(batch_root: Path, items: list[dict[str, Any]]) -> None:
    path = batch_root / "source_fix_queue.json"
    existing = read_json(path).get("items", []) if path.exists() else []
    merged = list(existing)
    seen = {(item.get("case_id", ""), item.get("segment_id", ""), tuple(item.get("source_turn_ids", []))) for item in merged if isinstance(item, dict)}
    for item in items:
        key = (item.get("case_id", ""), item.get("segment_id", ""), tuple(item.get("source_turn_ids", [])))
        if key in seen:
            continue
        merged.append(item)
        seen.add(key)
    write_json(path, {"schema_version": "source_fix_queue_v01", "items": merged})
def user_choice(row: dict[str, Any]) -> str:
    return str(row.get("人工结论") or row.get("your_choice") or "").strip()


def user_correction(row: dict[str, Any]) -> str:
    return str(row.get("人工修正") or row.get("corrected_value") or "").strip()


def user_notes(row: dict[str, Any]) -> str:
    return str(row.get("备注") or row.get("notes") or "").strip()


def apply_model_review_suggestions(segment: dict[str, Any]) -> None:
    review = segment.get("model_review", {}) if isinstance(segment.get("model_review", {}), dict) else {}
    issues = review.get("issues", []) if isinstance(review.get("issues", []), list) else []
    for issue in issues:
        if not isinstance(issue, dict):
            continue
        field = str(issue.get("field") or "")
        if field in CORRECTABLE_FIELDS:
            segment[field] = normalize_value(field, issue.get("suggested_value", ""))


def rows_from_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb["segments_review"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {str(headers[index]): value for index, value in enumerate(values)}
        rows.append(row)
    return rows


def apply_corrections(segment: dict[str, Any], row: dict[str, Any], corrected: str) -> bool:
    applied = False
    if corrected:
        try:
            patch = json.loads(corrected)
        except json.JSONDecodeError:
            segment["人工修正说明"] = corrected
            if str(segment.get("更优回复", "")).strip() == corrected.strip():
                segment["更优回复"] = str(row.get("主模型建议回复") or "")
            patch = {}
        if isinstance(patch, dict):
            for key, value in patch.items():
                if key in CORRECTABLE_FIELDS:
                    segment[key] = normalize_value(key, value)
                    applied = True
    return applied


def normalize_value(field: str, value: Any) -> Any:
    if field == "风险类型":
        if isinstance(value, list):
            return value
        return [item.strip() for item in str(value).split(",") if item.strip()]
    if field == "次要标签":
        if isinstance(value, dict):
            return value
        try:
            parsed = json.loads(str(value))
        except json.JSONDecodeError:
            return {"说明": str(value)}
        return parsed if isinstance(parsed, dict) else {"说明": str(value)}
    return value


def refresh_manifest(batch_root: Path) -> None:
    manifest_path = batch_root / "segments_manifest.json"
    if not manifest_path.exists():
        return
    manifest = read_json(manifest_path)
    for row in manifest.get("cases", []):
        case_path = batch_root / "cases" / row["case_id"] / "segments.json"
        if not case_path.exists():
            continue
        payload = read_json(case_path)
        segments = payload.get("segments", [])
        row["segment_count"] = len(segments)
        row["approved_count"] = sum(1 for item in segments if item.get("quality_status") == "approved")
        row["need_review_count"] = sum(1 for item in segments if item.get("need_human_review"))
        row["status"] = "ready" if row["approved_count"] else "needs_review"
    write_json(manifest_path, manifest)


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply human review choices to workV segments_v01.")
    parser.add_argument("--batch-id", required=True)
    parser.add_argument("--review-xlsx")
    args = parser.parse_args()
    print(json.dumps(apply_review(args.batch_id, args.review_xlsx), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
